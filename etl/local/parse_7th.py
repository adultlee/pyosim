"""7회(2018) 전국동시지방선거 읍면동별 개표결과 파서.

원본은 선거종류별 통합 xlsx 한 파일에 17개 시도가 모두 들어 있다.
구버전 파서의 치명 버그(파일 맨 위 후보 헤더를 한 번만 읽어 모든 시도에
서울 후보를 채움)를 피하기 위해, 블록(시도·선거구) 경계마다 나타나는
헤더 행에서 후보 컬럼→후보 매핑을 매번 갱신한다.

8회와 다른 7회만의 특징:
- 컬럼 라벨이 다르다. 시도지사·교육감은 [선거종류, 선거구명(=시도명), 시도명, 구시군명,
  읍면동명, …], 그 외(구시군장·시도의원·구시군의원·기초비례·교육의원)는 인덱스 1에
  '시도' 컬럼이 추가되고 광역비례는 '시도'는 있으나 '선거구명'이 없다.
  시도는 '시도' 컬럼이 있으면 그것을, 없으면 '선거구명'(값이 시도명)을 쓴다.
- 블록 헤더 행은 인덱스 0(선거종류) 셀이 채워진 행으로 식별한다.
- 후보 셀은 ``정당_x000D_\n이름`` 형태다. ``_x000D_``(반송 캐리지리턴 잔재)를 제거하고
  개행으로 정당/이름을 분리한다. 빈 후보 슬롯은 ``_x000D_`` 또는 빈칸이다.
- 구시군 합계행 라벨은 '계'(8회는 '합계'였다). 시도 합계행은 구시군명='합계'.
- 후보 컬럼 끝(유효표 '계')은 무효투표수 컬럼 바로 앞이다.
"""

import os
import unicodedata
from collections import Counter

import openpyxl

from etl.local.schema import normalize_level

# 파일명 (xxx) 키워드 → 선거종류
FILE_ETYPE = [
    ("시도지사", "시도지사"),
    ("구시군의장", "구시군장"),
    ("시도의회의원", "시도의원"),
    ("구시군의회의원", "구시군의원"),
    ("광역비례", "광역비례"),
    ("기초비례", "기초비례"),
    ("교육감", "교육감"),
    ("교육의원", "교육의원"),
]

# 후보 셀이 정당명만 있는(이름 없는) 비례대표 선거
PARTY_ONLY_TYPES = {"광역비례", "기초비례"}

# 구시군 단위 특수 투표(읍면동명 자리에 라벨이 들어옴)
SPECIAL_EUPMD = {"거소투표", "관외사전투표"}
# 읍면동 단위 실집계 구분
PRECINCT_GUBUN = {"관내사전투표", "선거일투표"}

# 구시군 합계행 라벨(읍면동명 자리). 7회는 '계'.
TOTAL_LABEL = "계"
# 읍면동 소계 구분(관내사전+선거일의 중복합 → 제외)
SUBTOTAL_GUBUN = "소계"
# 합계에 포함되는 정상적 '기타' 행(읍면동명 자리에 라벨). 합계 정합을 위해 emit 한다.
KNOWN_OTHER_EUPMD = {"잘못 투입·구분된 투표지", "잘못 투입·구분된 투표"}


def _clean(value):
    """셀 문자열에서 _x000D_(캐리지리턴 잔재)를 제거하고 양끝 공백을 정리."""
    if value is None:
        return ""
    return str(value).replace("_x000D_", "").strip()


def _to_int(value):
    """'129,816' 또는 정수 → 129816. 빈값·'·'·'-' → None."""
    if value is None:
        return None
    text = str(value).replace("_x000D_", "").replace(",", "").strip()
    if text in ("", "·", "-"):
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _column_index(header_row, *labels):
    """row0에서 라벨에 해당하는 컬럼 인덱스를 찾는다. 없으면 None."""
    for index, cell in enumerate(header_row):
        if _clean(cell) in labels:
            return index
    return None


def _parse_candidate_cell(cell, party_only):
    """후보 셀을 (정당, 후보자)로 해석. 빈 셀·'계'면 None 반환.

    - 비례: 정당명만 → (정당, None). 후보자는 없고 정당이 식별자다.
    - 정당_x000D_\\n후보자: _x000D_ 제거 후 split → (정당, 후보자).
    - 이름만(교육감·교육의원 무소속): (빈문자열, 이름).
    """
    text = _clean(cell)
    if text in ("", "계"):
        return None
    if party_only:
        return text, None
    parts = [piece.strip() for piece in text.split("\n") if piece.strip()]
    if len(parts) >= 2:
        return parts[0], parts[1]
    if len(parts) == 1:
        # 정당 없이 이름만(교육감·교육의원 무소속)
        return "", parts[0]
    return None


def _parse_file(filepath, election_type):
    """한 선거종류 통합 xlsx를 파싱해 (rows, totals, unrecognized)를 반환.

    unrecognized: 알려진 행 분류에 걸리지 않은 행들. emit 하지 않고 호출부가 경고.
    """
    workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        sheet_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header = list(sheet_rows[0])
    sido_col = _column_index(header, "시도")
    district_col = _column_index(header, "선거구명")
    eupmd_col = _column_index(header, "읍면동명")
    gubun_col = _column_index(header, "구분")
    electorate_col = _column_index(header, "선거인수")
    votes_col = _column_index(header, "투표수")
    invalid_col = _column_index(header, "무효투표수")
    abstain_col = _column_index(header, "기권수")

    # 시도는 '시도' 컬럼이 있으면 그것을, 없으면 '선거구명'(값=시도명)을 쓴다.
    province_col = sido_col if sido_col is not None else district_col
    # 시도지사·교육감은 선거구명이 곧 시도라 별도 선거구 접두가 필요 없다.
    has_district = district_col is not None and district_col != province_col

    # 필수 컬럼이 없으면(헤더 라벨 변경 등) 인덱스 산술이 모호한 TypeError로
    # 터지므로, 여기서 어느 라벨을 못 찾았는지 명확히 알린다.
    required = {
        "읍면동명": eupmd_col, "구분": gubun_col, "선거인수": electorate_col,
        "투표수": votes_col, "무효투표수": invalid_col,
    }
    missing = [label for label, index in required.items() if index is None]
    if missing or province_col is None:
        if province_col is None:
            missing.append("시도/선거구명")
        raise ValueError(f"{filepath}: 헤더에서 컬럼을 찾지 못함: {missing}")

    # 후보 컬럼 = 투표수 다음부터 유효표 '계'(무효투표수 바로 앞)까지.
    cand_start = votes_col + 1
    valid_col = invalid_col - 1  # '계'(유효표 합) 컬럼
    party_only = election_type in PARTY_ONLY_TYPES

    rows = []
    totals = []
    unrecognized = []
    cur_sido = cur_gu = cur_district = None
    col_to_candidate = {}

    for row_values in sheet_rows[1:]:
        row = list(row_values)

        if _clean(row[0]) != "":
            # 블록 헤더 행: 시도/선거구 갱신 + 후보 매핑 재구성
            cur_sido = _clean(row[province_col])
            cur_district = _clean(row[district_col]) if has_district else None
            col_to_candidate = {}
            for col in range(cand_start, valid_col):
                parsed = _parse_candidate_cell(row[col], party_only)
                if parsed is not None:
                    col_to_candidate[col] = parsed
            continue

        eupmd = _clean(row[eupmd_col])
        gubun = _clean(row[gubun_col])
        cur_gu = _clean(row[eupmd_col - 1])  # 구시군명은 읍면동명 바로 왼쪽

        if eupmd == "":
            continue  # 시도 합계행(구시군명='합계') 등: emit/합계 어디에도 넣지 않는다.

        turnout = _to_int(row[votes_col])

        if eupmd == TOTAL_LABEL:
            totals.append({
                "선거종류": election_type,
                "시도": cur_sido,
                "구시군": cur_gu,
                "level": None,
                "투표수": turnout,
            })
            continue

        if gubun == SUBTOTAL_GUBUN:
            continue  # 읍면동 소계는 관내사전+선거일의 합 → 중복이므로 제외

        if eupmd in SPECIAL_EUPMD:
            level = normalize_level(eupmd)
        elif gubun in PRECINCT_GUBUN:
            level = normalize_level(gubun)
        elif eupmd in KNOWN_OTHER_EUPMD:
            # '잘못 투입·구분된 투표지' 등 구시군 단위 기타. 합계에 포함되므로 emit.
            # normalize_level로 회차 간 라벨을 통일('잘못투입').
            level = normalize_level(eupmd)
        else:
            unrecognized.append({
                "선거종류": election_type,
                "시도": cur_sido,
                "구시군": cur_gu,
                "읍면동": eupmd,
                "구분": gubun,
            })
            continue

        # 선거인수는 선거일(당일)투표 행에서만 실제 선거인 수다. 그 외(사전·거소·관외·
        # 잘못투입)는 발급 매수 등이라 투표수보다 작을 수 있어 범위검사 오탐을 막으려 None.
        if level == "당일투표":
            electorate = _to_int(row[electorate_col])
        else:
            electorate = None

        # 한 구시군 안에 여러 선거구가 있어(시도의원·구시군의원 등) (읍면동, level) 키가
        # 겹친다. 선거구명을 읍면동에 접두해 같은 구시군 내에서 키를 유일하게 만든다.
        if has_district:
            precinct = f"{cur_district} {eupmd}"
        else:
            precinct = eupmd

        invalid = _to_int(row[invalid_col])
        abstain = _to_int(row[abstain_col])

        for col, (party, name) in col_to_candidate.items():
            rows.append({
                "선거_회차": 7,
                "선거종류": election_type,
                "시도": cur_sido,
                "구시군": cur_gu,
                "읍면동": precinct,
                "선거구명": cur_district,
                "선거인수": electorate,
                "투표수": turnout,
                "후보자": name,
                "정당": party,
                "득표수": _to_int(row[col]),
                "무효투표수": invalid,
                "기권수": abstain,
                "level": level,
            })

    return rows, totals, unrecognized


def parse_7th(dir_path):
    """7회 지방선거 디렉터리의 선거종류별 xlsx를 모두 파싱한다.

    반환: (rows, totals)
    - rows: 15개 스키마 컬럼의 tidy dict 리스트 (선거일은 오케스트레이터가 채움)
    - totals: 구시군 합계행 {선거종류,시도,구시군,level,투표수} 리스트. 한 구시군 안
      여러 선거구의 합계는 구시군 단위로 합산해 한 항목으로 낸다.
    """
    all_rows = []
    totals_by_key = {}
    all_unrecognized = []

    for filename in sorted(os.listdir(dir_path)):
        if not filename.endswith(".xlsx"):
            continue
        name = unicodedata.normalize("NFC", os.path.splitext(filename)[0])
        election_type = None
        for keyword, etype in FILE_ETYPE:
            if keyword in name:
                election_type = etype
                break
        if election_type is None:
            continue  # 국회의원선거(재보궐) 등은 제외

        file_rows, file_totals, file_unrecognized = _parse_file(
            os.path.join(dir_path, filename), election_type
        )
        all_rows.extend(file_rows)
        all_unrecognized.extend(file_unrecognized)
        for total in file_totals:
            key = (total["선거종류"], total["시도"], total["구시군"])
            if key not in totals_by_key:
                totals_by_key[key] = {
                    "선거종류": total["선거종류"],
                    "시도": total["시도"],
                    "구시군": total["구시군"],
                    "level": None,
                    "투표수": 0,
                }
            totals_by_key[key]["투표수"] += total["투표수"] or 0

    if all_unrecognized:
        _warn_unrecognized(all_unrecognized)

    totals = list(totals_by_key.values())
    return all_rows, totals


def _warn_unrecognized(unrecognized):
    """미인식 행을 (선거종류/시도/구시군/읍면동/구분)별로 집계해 경고 출력한다.

    크래시하지 않되, 조용히 사라지지 않도록 가시화한다. 정상이라면 0건이어야 한다.
    """
    counts = Counter(
        (row["선거종류"], row["시도"], row["구시군"], row["읍면동"], row["구분"])
        for row in unrecognized
    )
    print(
        f"[parse_7th] 경고: 미인식 행 {len(unrecognized)}건 "
        f"({len(counts)}종)을 emit 하지 않고 건너뜀:"
    )
    for (etype, sido, gu, eupmd, gubun), count in counts.most_common():
        print(
            f"  {count:6d}회  선거종류={etype} 시도={sido} 구시군={gu} "
            f"읍면동={eupmd!r} 구분={gubun!r}"
        )
