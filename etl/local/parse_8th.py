"""8회(2022) 전국동시지방선거 읍면동별 개표결과 파서.

원본은 선거종류별 통합 xlsx 한 파일에 17개 시도가 모두 들어 있다.
구버전 파서의 치명 버그(파일 맨 위 후보 헤더를 한 번만 읽어 모든 시도에
서울 후보를 채움)를 피하기 위해, 시도/구시군 블록 경계마다 나타나는
헤더 행에서 후보 컬럼→후보 매핑을 매번 갱신한다.

원본 레이아웃은 두 종류다.
- A형(시도지사·광역비례·교육감): [시도, 구시군, 읍면동, 구분, 선거인수, 투표수, 후보…, 계, 무효, 기권]
- B형(구시군장·시도의원·구시군의원·기초비례·교육의원): A형 + 인덱스 2에 선거구명 컬럼 추가
컬럼 위치는 row0의 헤더 라벨에서 동적으로 찾는다.

블록 경계 = 읍면동명이 빈 행. 그 행의 후보 컬럼에 정당/후보 식별자가 들어 있다.
"""

import os
import unicodedata
from collections import Counter

import openpyxl

from etl.local.schema import normalize_level

# 파일명 키워드 → 선거종류
FILE_ETYPE = [
    ("시도지사선거", "시도지사"),
    ("광역비례의원선거", "광역비례"),
    ("구시군장선거", "구시군장"),
    ("시도의원선거", "시도의원"),
    ("구시군의원선거", "구시군의원"),
    ("기초비례의원선거", "기초비례"),
    ("교육감선거", "교육감"),
    ("교육의원선거", "교육의원"),
]

# 후보 셀이 정당명만 있는(이름 없는) 비례대표 선거
PARTY_ONLY_TYPES = {"광역비례", "기초비례"}

# 구시군 단위 특수 투표(읍면동명 자리에 라벨이 들어옴)
SPECIAL_EUPMD = {"거소투표", "관외사전투표"}
# 읍면동 단위 실집계 구분
PRECINCT_GUBUN = {"관내사전투표", "선거일투표"}

# 구시군 합계행 라벨(읍면동명 자리)
TOTAL_LABEL = "합계"
# 읍면동 소계 구분(관내사전+선거일의 중복합 → 제외)
SUBTOTAL_GUBUN = "소계"
# 합계에 포함되는 정상적 '기타' 행(읍면동명 자리에 라벨). 합계 정합을 위해 emit 한다.
# 가운뎃점은 U+00B7(MIDDLE DOT). 원본 변형이 있을 수 있어 알려진 변형도 함께 둔다.
KNOWN_OTHER_EUPMD = {"잘못 투입·구분된 투표지", "잘못 투입·구분된 투표"}


def _to_int(value):
    """'129,816' → 129816. 빈값·'·'·'-' → None."""
    if value is None:
        return None
    text = str(value).replace(",", "").strip()
    if text in ("", "·", "-"):
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _column_index(header_row, *labels):
    """row0에서 라벨에 해당하는 컬럼 인덱스를 찾는다."""
    for index, cell in enumerate(header_row):
        text = str(cell or "").strip()
        if text in labels:
            return index
    raise ValueError(f"헤더에서 {labels} 컬럼을 찾지 못함")


def _parse_candidate_cell(cell, party_only):
    """후보 셀을 (정당, 후보자)로 해석. 빈 셀이면 None 반환.

    - 비례: 정당명만 → (정당, None). 후보자는 없고 정당이 식별자다.
    - 정당\n후보자: split → (정당, 후보자).
    - 이름만(교육감·교육의원 무소속): (빈문자열, 이름).
    """
    if not isinstance(cell, str):
        return None
    text = cell.strip()
    if text == "" or text == "\n":
        return None
    if party_only:
        return text, None
    party, separator, name = cell.partition("\n")
    party, name = party.strip(), name.strip()
    if separator == "":
        # 무소속 등 정당 없이 이름만
        return "", party
    if not name:
        return None
    return party, name


def _parse_file(filepath, election_type):
    """한 선거종류 통합 xlsx를 파싱해 (rows, totals, unrecognized)를 반환.

    unrecognized: 알려진 행 분류에 걸리지 않은 행들. emit 하지 않고 호출부가 경고.
    """
    workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        sheet_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header = list(sheet_rows[0])
    eupmd_col = _column_index(header, "읍면동명")
    gubun_col = _column_index(header, "구분")
    electorate_col = _column_index(header, "선거인수")
    votes_col = _column_index(header, "투표수")
    valid_col = _column_index(header, "계")
    invalid_col = _column_index(header, "무효투표수")
    abstain_col = _column_index(header, "기권수")
    cand_start = votes_col + 1
    # B형은 읍면동명이 인덱스 3 → 인덱스 2에 선거구명 컬럼이 있다.
    has_district = eupmd_col == 3
    district_col = 2 if has_district else None
    party_only = election_type in PARTY_ONLY_TYPES

    rows = []
    totals = []
    unrecognized = []
    cur_sido = cur_gu = cur_district = None
    col_to_candidate = {}

    for row_values in sheet_rows[2:]:
        row = list(row_values)
        eupmd = str(row[eupmd_col] or "").strip()
        gubun = str(row[gubun_col] or "").strip()

        if eupmd == "":
            # 블록 헤더 행: 시도/구시군/선거구 갱신 + 후보 매핑 재구성
            cur_sido = str(row[0] or "").strip()
            cur_gu = str(row[1] or "").strip()
            cur_district = str(row[district_col] or "").strip() if has_district else None
            col_to_candidate = {}
            # 후보 컬럼 = 투표수와 계 사이 전부
            for col in range(cand_start, valid_col):
                parsed = _parse_candidate_cell(row[col], party_only)
                if parsed is not None:
                    col_to_candidate[col] = parsed
            continue

        turnout = _to_int(row[votes_col])

        if eupmd == TOTAL_LABEL:
            totals.append({
                "선거종류": election_type,
                "시도": cur_sido,
                "구시군": cur_gu,
                "level": None,
                "투표수": turnout,
            })
            continue

        if gubun == SUBTOTAL_GUBUN:
            continue  # 읍면동 소계는 관내사전+선거일의 합 → 중복이므로 제외

        if eupmd in SPECIAL_EUPMD:
            level = normalize_level(eupmd)
        elif gubun in PRECINCT_GUBUN:
            level = normalize_level(gubun)
        elif eupmd in KNOWN_OTHER_EUPMD or gubun in KNOWN_OTHER_EUPMD:
            # '잘못 투입·구분된 투표지' 등 구시군 단위 기타 특수. 합계에 포함되므로 emit.
            # normalize_level로 회차 간 라벨을 통일('잘못투입').
            level = normalize_level(eupmd if eupmd in KNOWN_OTHER_EUPMD else gubun)
        else:
            # 미인식 행 shape(새 구분 라벨·개명된 특수투표·깨진 행 등). 조용히 데이터로
            # 흘려보내면 garbage level로 집계되므로, emit 하지 않고 수집해 끝에서 경고한다.
            unrecognized.append({
                "선거종류": election_type,
                "시도": cur_sido,
                "구시군": cur_gu,
                "읍면동": eupmd,
                "구분": gubun,
            })
            continue

        # 선거인수는 선거일(당일)투표 행에서만 실제 선거인 수다. 사전투표·거소·관외·
        # 잘못투입 행의 선거인수 칸은 선거인 정원이 아니라 발급 매수 등이라 투표수보다
        # 작을 수 있다(원본 그대로). 범위 검사(투표수≤선거인수) 오탐을 막으려 None 처리.
        if level == "당일투표":
            electorate = _to_int(row[electorate_col])
        else:
            electorate = None

        # B형은 한 구시군 안에 여러 선거구가 있어 (읍면동, level) 키가 겹친다.
        # 검증 게이트의 합계 검사는 구시군 단위로 (읍면동,level) 중복을 제거하므로,
        # 선거구명을 읍면동에 접두해 같은 구시군 내에서 키를 유일하게 만든다.
        if has_district:
            precinct = f"{cur_district} {eupmd}"
        else:
            precinct = eupmd

        invalid = _to_int(row[invalid_col])
        abstain = _to_int(row[abstain_col])

        for col, (party, name) in col_to_candidate.items():
            rows.append({
                "선거_회차": 8,
                "선거종류": election_type,
                "시도": cur_sido,
                "구시군": cur_gu,
                "읍면동": precinct,
                "선거구명": cur_district,
                "선거인수": electorate,
                "투표수": turnout,
                "후보자": name,
                "정당": party,
                "득표수": _to_int(row[col]),
                "무효투표수": invalid,
                "기권수": abstain,
                "level": level,
            })

    return rows, totals, unrecognized


def parse_8th(dir_path):
    """8회 지방선거 디렉터리의 선거종류별 xlsx를 모두 파싱한다.

    반환: (rows, totals)
    - rows: 15개 스키마 컬럼의 tidy dict 리스트 (선거일은 오케스트레이터가 채움)
    - totals: 구시군 합계행 {선거종류,시도,구시군,level,투표수} 리스트.
      검증 게이트는 구시군 단위로 합계를 검사하므로, 한 구시군 안 여러 선거구의
      합계는 구시군 단위로 합산해 한 항목으로 낸다.
    """
    all_rows = []
    totals_by_key = {}
    all_unrecognized = []

    for filename in sorted(os.listdir(dir_path)):
        if not filename.endswith(".xlsx"):
            continue
        name = unicodedata.normalize("NFC", os.path.splitext(filename)[0])
        election_type = None
        for keyword, etype in FILE_ETYPE:
            if keyword in name:
                election_type = etype
                break
        if election_type is None:
            continue  # 국회의원선거(재보궐) 등은 제외

        file_rows, file_totals, file_unrecognized = _parse_file(
            os.path.join(dir_path, filename), election_type
        )
        all_rows.extend(file_rows)
        all_unrecognized.extend(file_unrecognized)
        for total in file_totals:
            key = (total["선거종류"], total["시도"], total["구시군"])
            if key not in totals_by_key:
                totals_by_key[key] = {
                    "선거종류": total["선거종류"],
                    "시도": total["시도"],
                    "구시군": total["구시군"],
                    "level": None,
                    "투표수": 0,
                }
            totals_by_key[key]["투표수"] += total["투표수"] or 0

    if all_unrecognized:
        _warn_unrecognized(all_unrecognized)

    totals = list(totals_by_key.values())
    return all_rows, totals


def _warn_unrecognized(unrecognized):
    """미인식 행을 (선거종류/시도/구시군/읍면동/구분)별로 집계해 경고 출력한다.

    크래시하지 않되, 조용히 사라지지 않도록 가시화한다. 정상이라면 0건이어야 한다.
    """
    counts = Counter(
        (row["선거종류"], row["시도"], row["구시군"], row["읍면동"], row["구분"])
        for row in unrecognized
    )
    print(
        f"[parse_8th] 경고: 미인식 행 {len(unrecognized)}건 "
        f"({len(counts)}종)을 emit 하지 않고 건너뜀:"
    )
    for (etype, sido, gu, eupmd, gubun), count in counts.most_common():
        print(
            f"  {count:6d}회  선거종류={etype} 시도={sido} 구시군={gu} "
            f"읍면동={eupmd!r} 구분={gubun!r}"
        )
