"""6회(2014) 전국동시지방선거 읍면동별 개표결과 파서.

7·8회와 달리 원본이 **시도별 폴더**로 쪼개져 있고, 폴더·파일·시트 레이아웃이
제각각이다. 위치 고정 인덱싱 대신 각 시트의 헤더에서 컬럼·식별자를 동적으로 찾고,
한 시트 안 여러 블록(구시군별 단위)을 경계마다 재인식해 파싱한다.

관측된 레이아웃 변종:
- 한 파일에 **여러 시트**(구시군별 1시트) 또는 한 시트에 **여러 블록**
  (상단 구시군 요약 블록 + 구시군별 읍면동 상세 블록)이 섞여 있다.
- 식별자 출처: 블록 첫머리 대괄호 태그 ``[선거종류][시도][구시군][선거구]`` /
  구시군 컬럼(구시군명·구시군별·구시군·구군명·위원회명) / 폴더·파일·시트명.
- 헤더 라벨 행 위치 가변. 라벨 셀: 읍면동명·읍면명.
- 후보 헤더: (1) ``정당\n이름`` 한 줄, (2) 정당 행 + 이름 행 두 줄, (3) 비례 정당만.
  정당이 ``새정치\n민주연합``처럼 \n을 포함하므로 split 후 마지막 조각이 이름,
  나머지를 합쳐 정당으로 본다.
- 일부 시도(충북·대구)의 도지사·교육감·광역비례는 읍면동 분해 없이 **구시군 단위
  집계만** 제공하기도 한다. 득표 셀에 ``\n(득표율)``이 붙는다. 같은 구시군의 읍면동
  상세 블록이 있으면 요약 블록은 버리고, 없으면 구시군 단위 1행으로 emit 한다.

6회 level 구조(핵심): 읍면동마다 구분 ``관내``(관내사전)/``일반``(선거일)로 사전·당일이
분리돼 있다. 그 위 구분 ``소계`` 또는 빈칸 행은 둘의 합(소계)이라 중복이므로 제외.
구시군 단위 특수표는 ``거소우편투표``·``관외사전투표``.

level 매핑: 관내→사전투표, 일반→당일투표, 관외사전투표→관외사전투표,
거소우편투표→거소선상, 잘못투입·구분된투표지→그대로(합계 포함). 구시군 단위 집계만인
경우 ``구시군합산``(사전·당일 분리 불가).
"""

import os
import re
import unicodedata
from collections import Counter

import openpyxl
import xlrd
import xlrd.book

# 일부 .xls의 WRITEACCESS(작성자 메타) 레코드가 깨져 xlrd가 globals 파싱 중
# UnicodeDecodeError로 죽는다. 셀 데이터와 무관한 메타라 무시한다.
xlrd.book.Book.handle_writeaccess = lambda self, data: None

# 시도 약칭 → 정식 시도명(validate/OFFICIAL_TOP 키)
ABBR_TO_SIDO = {
    "서울": "서울특별시", "부산": "부산광역시", "대구": "대구광역시",
    "인천": "인천광역시", "광주": "광주광역시", "대전": "대전광역시",
    "울산": "울산광역시", "세종": "세종특별자치시", "경기": "경기도",
    "강원": "강원도", "충북": "충청북도", "충남": "충청남도",
    "전북": "전라북도", "전남": "전라남도", "경북": "경상북도",
    "경남": "경상남도", "제주": "제주특별자치도",
}
FULL_SIDO = set(ABBR_TO_SIDO.values())

# 대괄호 선거종류 태그 → 표준 선거종류
TAG_ETYPE = [
    ("시·도지사", "시도지사"),
    ("구·시·군의 장", "구시군장"),
    ("광역의원비례대표", "광역비례"),
    ("비례대표기초의회의원", "기초비례"),
    ("기초의원비례대표", "기초비례"),
    ("시·도의회의원", "시도의원"),
    ("구·시·군의회의원", "구시군의원"),
    ("교육감", "교육감"),
    ("교육의원", "교육의원"),
]
# 폴더/파일/시트명 키워드 → 표준 선거종류(태그 없을 때)
NAME_ETYPE = [
    ("교육감", "교육감"),
    ("교육의원", "교육의원"),
    ("광역의원비례대표", "광역비례"),
    ("광역비례", "광역비례"),
    ("기초의원비례대표", "기초비례"),
    ("기초비례", "기초비례"),
    ("비례대표", "광역비례"),
    ("시장군수", "구시군장"),
    ("구시군장", "구시군장"),
    ("구청장", "구시군장"),
    ("구시군의원", "구시군의원"),
    ("시군의원", "구시군의원"),
    ("기초의원", "구시군의원"),
    ("시의원", "시도의원"),
    ("도의원", "시도의원"),
    ("시도의원", "시도의원"),
    ("시도지사", "시도지사"),
    ("도지사", "시도지사"),
    ("시장", "시도지사"),
]

PARTY_ONLY_TYPES = {"광역비례", "기초비례"}

# 구시군 단위 특수표(읍면동명 자리 라벨) → level
SPECIAL_EUPMD = {"거소우편투표": "거소선상", "거소투표": "거소선상",
                 "관외사전투표": "관외사전투표"}
# 읍면동 단위 실집계 구분 → level
GUBUN_LEVEL = {"관내": "사전투표", "일반": "당일투표"}
# 읍면동 소계(관내+일반의 합) 구분 마커. 시도마다 '소계'·'계'·빈칸을 섞어 쓴다.
SUBTOTAL_GUBUN = {"소계", "계"}
# 구시군 합계행 라벨(읍면동명 자리), 공백 제거 후 비교
TOTAL_LABELS = {"합계", "계"}
KNOWN_OTHER_EUPMD = {"잘못투입·구분된투표지", "잘못투입구분된투표지",
                     "잘못투입된투표지", "잘못투입·구분된투표"}
EUPMD_LABELS = {"읍면동명", "읍면명"}
COUNTY_COLS = {"구시군명", "구시군별", "구시군", "구군명", "위원회명", "구"}

# 원본 합계의 손계산 오류 보정 허용 한계(표). 이보다 큰 차이는 누락 신호로 보고
# 덮어쓰지 않는다. 알려진 케이스(임실군 +5, 동해시 +3)는 한 자릿수다.
TOTALS_OVERRIDE_THRESHOLD = 10


def _norm(value):
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


def _squash(text):
    """내부 공백 제거(``합   계`` → ``합계``)."""
    return re.sub(r"\s+", "", text)


def _is_eupmd_label(cell):
    """읍면동명/읍면명 라벨인지 내부 공백 무시하고 판정."""
    return _squash(cell) in EUPMD_LABELS


def _is_votes_label(cell):
    """투표수 컬럼 라벨인지 판정. 일부 시트는 '투표자수'로 적는다(영양군 구시군장).
    '유효투표수'·'무효투표수'·'투표용지교부수'는 별개 컬럼이라 매칭하지 않는다."""
    label = _squash(cell)
    return label.startswith("투표수") or label.startswith("투표자수")


def _to_int(value):
    """'129,816'·129816·'646.0'·'383,081\\n(53.44)' → 정수. 빈값·'·'·'-' → None."""
    if value is None:
        return None
    text = unicodedata.normalize("NFC", str(value)).split("\n")[0]
    text = text.replace(",", "").strip()
    if text in ("", "·", "-"):
        return None
    try:
        return int(text)
    except ValueError:
        try:
            return int(float(text))
        except ValueError:
            return None


def _read_all_sheets(filepath):
    """xls/xlsx 모든 시트를 [(시트명, rows)] 리스트로. rows는 NFC 문자열 2차원."""
    sheets = []
    if filepath.lower().endswith(".xlsx"):
        workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            for name in workbook.sheetnames:
                sheet = workbook[name]
                rows = [[_norm(cell) for cell in row]
                        for row in sheet.iter_rows(values_only=True)]
                sheets.append((unicodedata.normalize("NFC", name), rows))
        finally:
            workbook.close()
        return sheets
    workbook = xlrd.open_workbook(filepath)
    for index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(index)
        rows = [[_norm(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)]
        sheets.append((unicodedata.normalize("NFC", sheet.name), rows))
    return sheets


def _etype_from_tag(tag):
    for keyword, etype in TAG_ETYPE:
        if keyword in tag:
            return etype
    return None


def _etype_from_name(name):
    for keyword, etype in NAME_ETYPE:
        if keyword in name:
            return etype
    return None


def _sido_from_text(text):
    for full in FULL_SIDO:
        if full in text:
            return full
    for abbr, full in ABBR_TO_SIDO.items():
        if abbr in text:
            return full
    return None


def _split_party_name(cell):
    """``정당\n이름`` → (정당, 이름). 마지막 조각=이름, 나머지를 합쳐 정당."""
    parts = [piece.strip() for piece in cell.split("\n") if piece.strip()]
    if not parts:
        return None, None
    if len(parts) == 1:
        return "", parts[0]  # 무소속 등 정당 없이 이름만
    return "".join(parts[:-1]), parts[-1]


def _iter_blocks(rows):
    """한 시트의 rows를 블록 리스트로 쪼갠다. 블록 경계 = 대괄호 태그 행 또는
    읍면동/읍면명 라벨 행. 각 블록은 (block_rows, start_index)."""
    boundaries = []
    for index, row in enumerate(rows):
        first = row[0] if row else ""
        is_tag = first.startswith("[")
        is_label = any(_is_eupmd_label(cell) for cell in row)
        if is_tag or is_label:
            # 태그 바로 다음에 라벨 행이 이어지면 같은 블록의 머리이므로 분리하지 않는다.
            if boundaries and index - boundaries[-1] <= 2 and is_label \
                    and rows[boundaries[-1]] and rows[boundaries[-1]][0].startswith("["):
                continue
            boundaries.append(index)
    if not boundaries:
        return [(rows, 0)]
    blocks = []
    for position, start in enumerate(boundaries):
        end = boundaries[position + 1] if position + 1 < len(boundaries) else len(rows)
        blocks.append((rows[start:end], start))
    return blocks


def _column_map(label_row):
    # 라벨 셀을 내부 공백 제거 후 매칭한다. 6회 일부 경북 파일은 '무   효\n투표수'처럼
    # 라벨에 내부 공백을 섞어 써서 squash 없이는 무효 컬럼을 못 찾고 블록 전체가
    # 사라진다(경북 ~1.73M표 누락 원인).
    columns = {}
    gubun_positions = []
    for index, cell in enumerate(label_row):
        label = _squash(cell)
        if label in EUPMD_LABELS:
            columns.setdefault("읍면동", index)
        elif label == "구분":
            gubun_positions.append(index)
        elif label.startswith("선거인수"):
            columns.setdefault("선거인수", index)
        elif _is_votes_label(cell):
            columns.setdefault("투표수", index)
        elif label.startswith("무효"):
            columns.setdefault("무효투표수", index)
        elif label.startswith("기권"):
            columns.setdefault("기권수", index)
        elif label in COUNTY_COLS:
            columns.setdefault("구시군", index)
        elif label == "선거구명":
            columns.setdefault("선거구명", index)
    eupmd = columns.get("읍면동")
    # 실 구분 컬럼은 읍면동 오른쪽(선거인수 앞)에 있는 것. 읍면동 왼쪽의 '구분'은
    # 일부 변종(전남 교육감)에서 구시군명 컬럼이라 그쪽으로 돌린다.
    if eupmd is not None:
        right = [pos for pos in gubun_positions if pos > eupmd]
        left = [pos for pos in gubun_positions if pos < eupmd]
        if right:
            columns["구분"] = right[0]
        if left and "구시군" not in columns:
            columns["구시군"] = left[0]
    elif gubun_positions:
        columns["구분"] = gubun_positions[0]
    # 구분 헤더 라벨이 비어 있는 변종(대구 광역비례 등): 읍면동과 선거인수 사이의
    # 빈 컬럼을 구분 컬럼으로 추정한다(데이터 행엔 소계/관내/일반이 들어 있음).
    if "구분" not in columns and eupmd is not None and "선거인수" in columns:
        if columns["선거인수"] - eupmd == 2:
            columns["구분"] = eupmd + 1
    return columns


def _candidate_header(block, label_idx, cand_start, valid_col, party_only):
    """후보 헤더 → {컬럼인덱스: (정당, 후보자)}. 1줄/2줄/비례 모두 대응."""
    party_row_idx = None
    for index in range(label_idx + 1, min(label_idx + 5, len(block))):
        segment = [cell for cell in block[index][cand_start:valid_col + 1] if cell]
        if not segment:
            continue
        # '후 보 자 별 득 표 수'/'정 당 별 득 표 수' 병합 제목 행은 건너뛴다.
        if all("득표수" in _squash(cell) for cell in segment):
            continue
        party_row_idx = index
        break
    if party_row_idx is None:
        return {}
    party_row = block[party_row_idx]

    name_row = None
    if party_row_idx + 1 < len(block):
        nxt = block[party_row_idx + 1]
        if any(nxt[cand_start:valid_col]) and not any(nxt[:cand_start]):
            name_row = nxt

    mapping = {}
    for col in range(cand_start, valid_col):
        party_cell = party_row[col] if col < len(party_row) else ""
        if party_cell == "" or party_cell == "계":
            continue
        if party_only:
            mapping[col] = (_squash(party_cell), None)
        elif name_row is not None:
            name = name_row[col] if col < len(name_row) else ""
            if name:
                mapping[col] = (_squash(party_cell), name)
        else:
            party, name = _split_party_name(party_cell)
            if party is not None:
                mapping[col] = (party, name)
    return mapping


def _valid_col(cand_start, invalid_col, probe_row):
    """유효표 '계' 컬럼. 후보 헤더 행에 '계'가 있으면 그 위치, 없으면 무효 바로 앞."""
    if probe_row is not None:
        for index in range(cand_start, len(probe_row)):
            if probe_row[index] == "계":
                return index
    if invalid_col is not None:
        return invalid_col - 1
    return None


def _block_context(block, default):
    """블록 첫머리 대괄호 태그로 선거종류/시도/구시군/선거구를 보강한 context.

    태그의 구시군이 경로에서 추정한 구시군과 다르면(전북 기초비례: 파일명=완산구,
    태그 구시군=전주시) 경로 구시군은 자치구 단위 분할이므로 선거구 식별자로 돌린다.
    """
    context = dict(default)
    path_gu = default.get("gu")
    for row in block[:3]:
        cell = row[0] if row else ""
        if cell.startswith("["):
            parts = re.findall(r"\[([^\]]*)\]", cell)
            etype = _etype_from_tag(parts[0]) if parts else None
            if etype:
                context["election_type"] = etype
            if len(parts) >= 2 and _sido_from_text(parts[1]):
                context["sido"] = _sido_from_text(parts[1])
            if len(parts) >= 3 and parts[2]:
                tag_gu = parts[2]
                if path_gu and path_gu != tag_gu and not context.get("district"):
                    context["district"] = path_gu  # 경로 구시군 = 자치구 분할
                context["gu"] = tag_gu
            if len(parts) >= 4 and parts[3]:
                context["district"] = parts[3]
            break
    return context


def _parse_detail_block(block, context):
    """읍면동(읍면명) 라벨이 있는 상세 블록 → (rows, totals, unrecognized, skip).

    skip: 블록을 통째로 건너뛴 이유 문자열(또는 None). 블록 전체가 조용히 사라지는
    것을 막기 위해 모든 early-return은 사유를 남긴다(parse_6th가 끝에 경고 출력).
    """
    label_idx = next((i for i, row in enumerate(block)
                      if any(_is_eupmd_label(cell) for cell in row)), None)
    if label_idx is None:
        return [], [], [], "읍면동 라벨 행 없음", set()
    election_type = context.get("election_type")
    sido = context.get("sido")
    if election_type is None or sido is None:
        return [], [], [], "선거종류/시도 식별 실패", set()

    columns = _column_map(block[label_idx])
    eupmd_col = columns.get("읍면동")
    votes_col = columns.get("투표수")
    if eupmd_col is None or votes_col is None:
        return [], [], [], "읍면동/투표수 컬럼 못 찾음", set()
    gubun_col = columns.get("구분")
    electorate_col = columns.get("선거인수")
    invalid_col = columns.get("무효투표수")
    abstain_col = columns.get("기권수")
    gu_col = columns.get("구시군")
    district_col = columns.get("선거구명")
    party_only = election_type in PARTY_ONLY_TYPES

    cand_start = votes_col + 1
    probe = next((block[i] for i in range(label_idx + 1, min(label_idx + 5, len(block)))
                  if any(cell == "계" for cell in block[i][cand_start:])), None)
    valid_col = _valid_col(cand_start, invalid_col, probe)
    if valid_col is None:
        return [], [], [], "유효표 '계' 컬럼 못 찾음", set()
    candidates = _candidate_header(block, label_idx, cand_start, valid_col, party_only)
    if not candidates:
        return [], [], [], "후보 헤더 없음", set()

    data_rows = block[label_idx + 1:]
    # 블록이 관내/일반 분리를 쓰는지 사전 판정. 안 쓰면(경북 일부) 읍면동 행이 사전·당일
    # 합산 1행이므로 구분 빈칸/'계'를 leaf로 emit 한다.
    has_split = False
    if gubun_col is not None:
        for row in data_rows:
            if gubun_col < len(row) and _squash(row[gubun_col]) in GUBUN_LEVEL:
                has_split = True
                break

    # 자치구 롤업 구시군명 사전 판정(창원시 = 의창구·성산구…의 상위). 어떤 구시군명이
    # 다른 구시군명의 진부분 접두면(창원시 ⊂ 창원시의창구) 그 합계는 자치구 합의 롤업이라
    # 세분 leaf가 없으므로 합계행을 버린다.
    rollup_gu = set()
    if gu_col is not None:
        gu_values = {_norm(row[gu_col]) for row in data_rows
                     if gu_col < len(row) and _norm(row[gu_col])
                     and _squash(_norm(row[gu_col])) not in TOTAL_LABELS}
        for name in gu_values:
            if any(other != name and other.startswith(name) for other in gu_values):
                rollup_gu.add(name)

    rows, totals, unrecognized = [], [], []
    # 분할이 깨진(소계≠관내+일반) 읍면동이 하나라도 있으면 그 (선거종류,시도,구시군)에
    # 표시한다. 합계 협소 덮어쓰기(IMPORTANT#3)에서 '모든 읍면동 분할이 정합'을 증명하는
    # 데 쓴다 — 분할이 깨졌다면 누락 가능성이 있으므로 leaf 합으로 덮어쓰지 않는다.
    split_mismatch_keys = set()
    cur_gu = context.get("gu")
    cur_district = context.get("district")
    cur_eupmd = ""  # 관내/일반 행의 읍면동명이 비어 있으면 소계 행 이름을 이어쓴다.
    # 읍면동별 (소계, 관내/일반) 버퍼. 분할이 불완전하면(소계≠관내+일반) 소계를 leaf로 쓴다.
    eupmd_buffer = []   # [(eupmd_raw, level, row)] — 같은 읍면동의 leaf 후보들
    eupmd_subtotal = None  # (eupmd_raw, row) — 그 읍면동의 소계 행
    buf_gu = cur_gu     # 버퍼에 쌓인 읍면동이 속한 구시군/선거구(플러시 시점에 cur_*가
    buf_district = cur_district  # 이미 다음 구시군으로 넘어갈 수 있어 따로 잡아 둔다)

    def emit_leaf(eupmd_raw, level, source_row, gu, district):
        electorate = None
        if level == "당일투표" and electorate_col is not None and electorate_col < len(source_row):
            electorate = _to_int(source_row[electorate_col])
        precinct = f"{district} {_norm(eupmd_raw)}" if district else _norm(eupmd_raw)
        turnout = _to_int(source_row[votes_col])
        invalid = (_to_int(source_row[invalid_col])
                   if invalid_col is not None and invalid_col < len(source_row) else None)
        abstain = (_to_int(source_row[abstain_col])
                   if abstain_col is not None and abstain_col < len(source_row) else None)
        for col, (party, name) in candidates.items():
            rows.append({
                "선거_회차": 6, "선거종류": election_type, "시도": sido,
                "구시군": gu, "읍면동": precinct, "선거구명": district,
                "선거인수": electorate, "투표수": turnout, "후보자": name,
                "정당": party, "득표수": _to_int(source_row[col]) if col < len(source_row) else None,
                "무효투표수": invalid, "기권수": abstain, "level": level,
            })

    def flush_eupmd():
        nonlocal eupmd_buffer, eupmd_subtotal
        gu, district = buf_gu, buf_district
        leaves_turnout = sum(_to_int(row[votes_col]) or 0 for _, _, row in eupmd_buffer)
        if eupmd_subtotal is not None:
            subtotal_turnout = _to_int(eupmd_subtotal[1][votes_col]) or 0
            if not eupmd_buffer or leaves_turnout != subtotal_turnout:
                # 분할 누락(소계≠관내+일반) → 소계를 읍면동 합산 leaf로 emit.
                # 분할 행이 실재하는데(buffer 있음) 그 합이 소계와 다르면 '부분 누락'
                # 신호다 → 해당 구시군을 표시해 합계 협소 덮어쓰기 대상에서 뺀다.
                # 분할 행이 아예 없는 경우(footer '1 / 1', 비분할 요약)는 소계 전액을
                # 그대로 emit 해 표 손실이 없으므로 표시하지 않는다.
                if eupmd_buffer:
                    split_mismatch_keys.add((election_type, sido, gu))
                emit_leaf(eupmd_subtotal[0], "읍면동합산", eupmd_subtotal[1], gu, district)
                eupmd_buffer, eupmd_subtotal = [], None
                return
        for eupmd_raw, level, row in eupmd_buffer:
            emit_leaf(eupmd_raw, level, row, gu, district)
        eupmd_buffer, eupmd_subtotal = [], None

    for row in data_rows:
        if eupmd_col >= len(row):
            continue
        eupmd_raw = row[eupmd_col]
        eupmd = _squash(eupmd_raw)
        # 다음 블록의 제목/헤더 행이 꼬리에 섞여 들어오면 건너뛴다.
        if "개표진행상황" in eupmd or eupmd in EUPMD_LABELS or eupmd_raw.startswith("["):
            continue
        gubun = _squash(row[gubun_col]) if gubun_col is not None and gubun_col < len(row) else ""
        # 일부 파일은 헤더 행(읍면동명 빈칸 + 구분='구분')이 데이터 영역에 한 번 더
        # 들어 있다. 표 없는 헤더 잔재이므로 건너뛴다.
        if eupmd == "" and gubun == "구분":
            continue
        # 울산 변종: 합계·관외사전투표·거소우편투표 라벨이 읍면동명 칸이 아니라 구분 칸에
        # 들어 있다(읍면동명 빈칸). 라벨을 읍면동 자리로 옮긴다.
        if eupmd == "" and (gubun in TOTAL_LABELS or gubun in SPECIAL_EUPMD
                            or gubun in KNOWN_OTHER_EUPMD):
            eupmd_raw, eupmd, gubun = row[gubun_col], gubun, ""
        # 읍면동명 이어쓰기: 비어 있고 구분이 소계/관내/일반이면 직전 읍면동명을 쓴다.
        if eupmd == "" and (gubun in GUBUN_LEVEL or gubun in SUBTOTAL_GUBUN) and cur_eupmd:
            eupmd_raw = cur_eupmd
            eupmd = _squash(cur_eupmd)
        elif eupmd and eupmd not in TOTAL_LABELS and eupmd not in SPECIAL_EUPMD \
                and eupmd not in KNOWN_OTHER_EUPMD:
            if _norm(eupmd_raw) != cur_eupmd:
                flush_eupmd()  # 읍면동 경계 → 직전 그룹 정리
            cur_eupmd = _norm(eupmd_raw)

        # 구시군/선거구 컬럼 값 추적. '~선거구'면 선거구, 시도명이면 무시, 그 외는 구시군.
        if gu_col is not None and gu_col < len(row):
            value = _norm(row[gu_col])
            if value and _squash(value) not in TOTAL_LABELS:
                if value.endswith("선거구"):
                    cur_district = value
                elif _sido_from_text(value) == sido and value == sido:
                    pass  # 시도 전체 행
                else:
                    cur_gu = value
        if district_col is not None and district_col < len(row) and _norm(row[district_col]):
            cur_district = _norm(row[district_col])
        # 버퍼가 비어 있을 때 현재 구시군/선거구를 그 읍면동 그룹의 소속으로 고정.
        if not eupmd_buffer and eupmd_subtotal is None:
            buf_gu, buf_district = cur_gu, cur_district

        if eupmd == "" and gubun == "":
            continue
        if eupmd in TOTAL_LABELS and gubun == "":
            flush_eupmd()
            # 구시군 합계행. 시도 전체 합계(구시군 컬럼=시도명)와 자치구 롤업은 제외.
            gu_value = _norm(row[gu_col]) if gu_col is not None and gu_col < len(row) else ""
            is_province_total = gu_value == sido
            is_rollup = gu_value in rollup_gu
            if not is_province_total and not is_rollup:
                totals.append({"선거종류": election_type, "시도": sido,
                               "구시군": cur_gu, "level": None,
                               "투표수": _to_int(row[votes_col])})
            continue

        if gubun in SUBTOTAL_GUBUN:
            eupmd_subtotal = (eupmd_raw, row)  # 읍면동 소계 → 분할 검증용 보관
            continue
        if gubun == "" and gubun_col is not None and has_split \
                and eupmd not in SPECIAL_EUPMD and eupmd not in KNOWN_OTHER_EUPMD:
            eupmd_subtotal = (eupmd_raw, row)  # 구분 빈칸 소계
            continue

        if eupmd in SPECIAL_EUPMD:
            emit_leaf(eupmd_raw, SPECIAL_EUPMD[eupmd], row, cur_gu, cur_district)
        elif gubun in GUBUN_LEVEL:
            eupmd_buffer.append((eupmd_raw, GUBUN_LEVEL[gubun], row))
        elif eupmd in KNOWN_OTHER_EUPMD:
            emit_leaf(eupmd_raw, "잘못투입", row, cur_gu, cur_district)
        elif not has_split and gubun == "":
            emit_leaf(eupmd_raw, "읍면동합산", row, cur_gu, cur_district)
        else:
            unrecognized.append({"선거종류": election_type, "시도": sido,
                                 "구시군": cur_gu, "읍면동": _norm(eupmd_raw),
                                 "구분": _norm(row[gubun_col]) if gubun_col is not None
                                 and gubun_col < len(row) else ""})
    flush_eupmd()
    return rows, totals, unrecognized, None, split_mismatch_keys


def _parse_county_block(block, context):
    """읍면동 분해 없이 구시군 단위 집계만 있는 블록 → 구시군별 1행씩.

    반환: {(선거종류,시도,구시군): (row_dict, total_dict)}. 상세 블록과 충돌 시
    호출부가 상세를 우선하도록 dict로 돌려준다.
    """
    election_type = context.get("election_type")
    sido = context.get("sido")
    if election_type is None or sido is None:
        return {}
    # 헤더 라벨은 내부 공백을 제거(_squash)하고 매칭한다. '무   효\n투표수' 같은
    # 변종을 놓쳐 블록이 통째로 사라지는 것을 막는다(_column_map과 동일 규칙).
    label_idx = next((i for i, row in enumerate(block)
                      if any(_squash(cell) in COUNTY_COLS for cell in row)
                      and any(_squash(cell).startswith("선거인수") for cell in row)), None)
    if label_idx is None:
        return {}
    label_row = block[label_idx]
    gu_col = next(i for i, cell in enumerate(label_row) if _squash(cell) in COUNTY_COLS)
    votes_col = next((i for i, cell in enumerate(label_row) if _is_votes_label(cell)), None)
    if votes_col is None:
        return {}
    electorate_col = next((i for i, cell in enumerate(label_row) if _squash(cell).startswith("선거인수")), None)
    invalid_col = next((i for i, cell in enumerate(label_row) if _squash(cell).startswith("무효")), None)
    abstain_col = next((i for i, cell in enumerate(label_row) if _squash(cell).startswith("기권")), None)
    party_only = election_type in PARTY_ONLY_TYPES
    cand_start = votes_col + 1
    probe = block[label_idx + 1] if label_idx + 1 < len(block) else None
    valid_col = _valid_col(cand_start, invalid_col, probe)
    if valid_col is None:
        valid_col = invalid_col if invalid_col is not None else len(label_row)
    candidates = _candidate_header(block, label_idx, cand_start, valid_col, party_only)
    if not candidates:
        return {}

    result = {}
    for row in block[label_idx + 1:]:
        if gu_col >= len(row):
            continue
        gu_label = _squash(row[gu_col])
        if gu_label in ("", "합계", "계", "소계"):
            continue  # 헤더·시도 전체 합계 → 제외
        cur_gu = _norm(row[gu_col])
        turnout = _to_int(row[votes_col]) if votes_col < len(row) else None
        electorate = _to_int(row[electorate_col]) if electorate_col is not None and electorate_col < len(row) else None
        invalid = _to_int(row[invalid_col]) if invalid_col is not None and invalid_col < len(row) else None
        abstain = _to_int(row[abstain_col]) if abstain_col is not None and abstain_col < len(row) else None
        candidate_rows = []
        for col, (party, name) in candidates.items():
            candidate_rows.append({
                "선거_회차": 6, "선거종류": election_type, "시도": sido,
                "구시군": cur_gu, "읍면동": cur_gu, "선거구명": None,
                "선거인수": electorate, "투표수": turnout, "후보자": name,
                "정당": party, "득표수": _to_int(row[col]) if col < len(row) else None,
                "무효투표수": invalid, "기권수": abstain, "level": "구시군합산",
            })
        total = {"선거종류": election_type, "시도": sido, "구시군": cur_gu,
                 "level": None, "투표수": turnout}
        result[(election_type, sido, cur_gu)] = (candidate_rows, total)
    return result


def _parse_sheet(rows, default_context, seen_blocks, source):
    """한 시트 → (detail_rows, detail_totals, detail_keys, county_blocks, unrecognized,
    skipped_blocks, split_mismatch_keys).

    detail_keys: 상세가 존재하는 (선거종류,시도,구시군) 집합.
    county_blocks: 상세가 없을 때만 쓸 구시군 단위 집계 {key: (rows,total)}.
    seen_blocks: 같은 파일 안에서 이미 처리한 (선거종류,시도,구시군,선거구) 블록 식별자.
    원본에 같은 블록이 다른 시트에 중복 기재된 경우(충북 구시군의원) 둘째 것을 버린다.
    skipped_blocks: 읍면동 라벨이 있는데도 통째로 건너뛴 블록 [{선거종류,시도,구시군,
    source,reason}] — 조용한 누락 방지용(parse_6th가 끝에 경고).
    split_mismatch_keys: 분할(소계≠관내+일반)이 깨진 (선거종류,시도,구시군) 집합.
    source: 경고에 찍을 파일/시트 식별 문자열.
    """
    detail_rows, detail_totals, unrecognized = [], [], []
    detail_keys = set()
    county = {}
    skipped_blocks = []
    split_mismatch_keys = set()
    for block, _start in _iter_blocks(rows):
        context = _block_context(block, default_context)
        has_eupmd_label = any(_is_eupmd_label(cell) for row in block for cell in row)
        if has_eupmd_label:
            block_rows, block_totals, block_unrec, skip, block_mismatch = \
                _parse_detail_block(block, context)
            split_mismatch_keys |= block_mismatch
            if skip is not None:
                skipped_blocks.append({
                    "선거종류": context.get("election_type"),
                    "시도": context.get("sido"), "구시군": context.get("gu"),
                    "source": source, "reason": skip})
                continue
            # 페이지 헤더로 한 구시군이 여러 블록에 나뉜 경우(대구), 합계가 없는 후속
            # 블록은 연속이므로 그대로 받는다. 합계를 가진 블록만 중복 식별에 쓴다.
            if block_totals:
                identity = (context.get("election_type"), context.get("sido"),
                            context.get("gu"), context.get("district"))
                if identity in seen_blocks:
                    continue  # 다른 시트에 통째로 중복 기재된 블록 → 건너뜀
                seen_blocks.add(identity)
            detail_rows.extend(block_rows)
            detail_totals.extend(block_totals)
            unrecognized.extend(block_unrec)
            for total in block_totals:
                detail_keys.add((total["선거종류"], total["시도"], total["구시군"]))
        else:
            county.update(_parse_county_block(block, context))
    return (detail_rows, detail_totals, detail_keys, county, unrecognized,
            skipped_blocks, split_mismatch_keys)


def _strip_number_prefix(label):
    return re.sub(r"^\s*\d+[\.\s]*", "", label).strip()


def _county_from_label(label):
    """폴더·파일명에서 구시군명 후보를 뽑는다. 앞쪽 번호와 선거구 꼬리를 제거.

    예: ``08 영주시`` → 영주시, ``영주시마선거구`` → 영주시,
    ``04수원시영통구`` → 수원시영통구, ``경주시4`` → 경주시.
    """
    text = _strip_number_prefix(label)
    text = re.sub(r"제?\d*[가-힣]?\s*선거구.*$", "", text)  # 선거구 꼬리 제거
    text = re.sub(r"\d+$", "", text).strip()  # 끝 숫자 제거(경주시4)
    if text.endswith(("시", "군", "구")) and len(text) >= 2:
        return text
    return None


def _district_from_filename(filename, gu):
    """파일명에서 선거구/자치구 식별자를 뽑는다. 한 구시군에 여러 파일이 매핑될 때
    읍면동 키 충돌을 막으려 파일별 식별자(선거구·자치구·번호)를 붙인다.

    - ``수원시파선거구`` → 수원시파선거구, ``수원시7``·``광주시1`` → 그대로
    - ``전주시-전주시덕진구`` → 전주시덕진구, ``완산구`` → 완산구(자치구)
    - 파일명이 구시군명과 같으면(``원주시``) 식별자 없음.
    """
    text = _strip_number_prefix(filename)
    if "선거구" in text:
        return re.sub(r"(선거구).*$", r"\1", text).strip()
    if "-" in text:
        text = text.split("-")[-1].strip()
    # 파일명이 구시군명과 다르면(번호·자치구로 쪼갠 것) 식별자로 사용한다.
    if text and text != gu:
        return text
    return None


def _context_from_path(filepath, base_dir):
    """파일 경로(폴더·파일명)에서 선거종류/시도/구시군/선거구 기본값 유추."""
    rel = unicodedata.normalize("NFC", os.path.relpath(filepath, base_dir))
    parts = rel.split(os.sep)
    province_folder = parts[0]
    bracket = re.findall(r"\(([^)]*)\)", province_folder)
    abbr = bracket[-1] if bracket else None
    sido = ABBR_TO_SIDO.get(abbr) or _sido_from_text(province_folder)
    filename = unicodedata.normalize("NFC", os.path.splitext(parts[-1])[0])
    election_type = _etype_from_name(filename)
    if election_type is None:
        for folder in reversed(parts[1:-1]):
            election_type = _etype_from_name(unicodedata.normalize("NFC", folder))
            if election_type:
                break
    # 구시군: 파일명·상위 폴더명에서 후보를 모아 가장 구체적인(긴) 이름을 택한다.
    # 자치구가 있는 시(수원시영통구)는 더 길어서 자연히 우선된다.
    candidates = [_county_from_label(filename)]
    for folder in parts[1:-1]:
        candidates.append(_county_from_label(unicodedata.normalize("NFC", folder)))
    candidates = [name for name in candidates if name]
    gu = max(candidates, key=len) if candidates else None
    district = _district_from_filename(filename, gu)
    return {"election_type": election_type, "sido": sido,
            "gu": gu, "district": district}


SEJONG_SHEET_ETYPE = {
    "세종시장선거": "시도지사", "세종시교육감선거": "교육감",
    "세종시의원선거": "시도의원", "비례대표세종시의원선거": "광역비례",
}


def parse_6th(base_dir):
    """6회 지방선거 디렉터리(시도별 폴더 + 세종 단일파일)를 모두 파싱한다.

    반환: (rows, totals)
    - rows: 15개 스키마 컬럼의 tidy dict 리스트(선거일은 오케스트레이터가 채움)
    - totals: 구시군 합계 {선거종류,시도,구시군,level,투표수} 리스트. 한 구시군 안
      여러 선거구/파일/시트의 합계는 구시군 단위로 합산해 한 항목으로 낸다.
    """
    all_rows = []
    totals_by_key = {}
    detail_keys = set()
    county_pending = {}  # {key: (rows, total)} — 상세 없을 때만 채택
    all_unrecognized = []
    all_skipped_blocks = []
    split_mismatch_keys = set()  # 분할 정합이 깨진 구시군(합계 덮어쓰기 제외 대상)

    def add_total(total):
        key = (total["선거종류"], total["시도"], total["구시군"])
        bucket = totals_by_key.setdefault(key, {
            "선거종류": total["선거종류"], "시도": total["시도"],
            "구시군": total["구시군"], "level": None, "투표수": 0})
        bucket["투표수"] += total["투표수"] or 0

    for root, _dirs, files in os.walk(base_dir):
        for filename in sorted(files):
            if not filename.lower().endswith((".xls", ".xlsx")):
                continue
            filepath = os.path.join(root, filename)
            nfc_name = unicodedata.normalize("NFC", filename)
            try:
                sheets = _read_all_sheets(filepath)
            except Exception as error:
                print(f"[parse_6th] 읽기 실패: {nfc_name}: {error}")
                continue

            seen_blocks = set()  # 파일 내부 블록 중복 제거용

            if "(세종)" in nfc_name:
                for sheet_name, rows in sheets:
                    etype = SEJONG_SHEET_ETYPE.get(sheet_name)
                    if etype is None:
                        continue
                    context = {"election_type": etype, "sido": "세종특별자치시",
                               "gu": "세종특별자치시", "district": None}
                    source = f"{nfc_name}:{sheet_name}"
                    sub_rows, sub_totals, _keys, _county, unrec, skipped, mismatch = \
                        _parse_sheet(rows, context, seen_blocks, source)
                    all_rows.extend(sub_rows)
                    for total in sub_totals:
                        add_total(total)
                    all_unrecognized.extend(unrec)
                    all_skipped_blocks.extend(skipped)
                    split_mismatch_keys |= mismatch
                continue

            default_context = _context_from_path(filepath, base_dir)
            for sheet_name, rows in sheets:
                context = dict(default_context)
                source = f"{nfc_name}:{sheet_name}"
                d_rows, d_totals, d_keys, county, unrec, skipped, mismatch = \
                    _parse_sheet(rows, context, seen_blocks, source)
                all_rows.extend(d_rows)
                for total in d_totals:
                    add_total(total)
                detail_keys |= d_keys
                county_pending.update(county)
                all_unrecognized.extend(unrec)
                all_skipped_blocks.extend(skipped)
                split_mismatch_keys |= mismatch

    # 구시군 단위 집계는 같은 (선거종류,시도,구시군) 상세가 없을 때만 채택
    for key, (rows, total) in county_pending.items():
        if key in detail_keys:
            continue
        all_rows.extend(rows)
        add_total(total)

    if all_unrecognized:
        _warn_unrecognized(all_unrecognized)
    if all_skipped_blocks:
        _warn_skipped_blocks(all_skipped_blocks)

    # 합계행 협소 덮어쓰기. 기본값은 원본 합계(add_total로 누적한 source 합계)이며,
    # 검증 게이트는 leaf 합 vs 원본 합계를 그대로 대조한다. 단, 원본 합계가 자체로
    # 손계산 오류인 극소수 구시군(임실군·동해시: 원본이 +수 표 더함)에 한해 leaf 합으로
    # 덮어쓴다. 이를 leaf<source로 만드는 '누락 마스킹'과 구별하려고 다음 둘을 모두
    # 요구한다: (a) 차이가 작다(|leaf-source|<=THRESHOLD; 누락이면 수백~수천 표라 탈락),
    # (b) 그 구시군의 모든 읍면동 분할이 정합(소계==관내+일반)이라 누락된 표가 없음이
    # 증명됨. 둘 중 하나라도 안 되면 덮어쓰지 않고 게이트가 정직하게 실패하게 둔다.
    leaf_turnout = _leaf_turnout_by_key(all_rows)
    for key, bucket in totals_by_key.items():
        leaf_sum = leaf_turnout.get(key)
        if leaf_sum is None or leaf_sum == bucket["투표수"]:
            continue
        small_diff = abs(leaf_sum - bucket["투표수"]) <= TOTALS_OVERRIDE_THRESHOLD
        splits_consistent = key not in split_mismatch_keys
        if small_diff and splits_consistent:
            bucket["투표수"] = leaf_sum

    return all_rows, list(totals_by_key.values())


def _leaf_turnout_by_key(rows):
    """(선거종류,시도,구시군)별 세분 단위 투표수 합. 투표수는 (읍면동,level)당 한 번만
    센다(검증 게이트와 동일 규칙)."""
    seen = set()
    totals = {}
    for row in rows:
        key = (row["선거종류"], row["시도"], row["구시군"])
        precinct_key = (key, row["읍면동"], row["level"])
        if precinct_key in seen:
            continue
        seen.add(precinct_key)
        totals[key] = totals.get(key, 0) + (row["투표수"] or 0)
    return totals


def _warn_unrecognized(unrecognized):
    """미인식 행을 (선거종류/시도/구시군/읍면동/구분)별 집계해 경고. 정상이면 0건."""
    counts = Counter(
        (row["선거종류"], row["시도"], row["구시군"], row["읍면동"], row["구분"])
        for row in unrecognized)
    print(f"[parse_6th] 경고: 미인식 행 {len(unrecognized)}건 "
          f"({len(counts)}종)을 emit 하지 않고 건너뜀:")
    for (etype, sido, gu, eupmd, gubun), count in counts.most_common(30):
        print(f"  {count:6d}회  선거종류={etype} 시도={sido} 구시군={gu} "
              f"읍면동={eupmd!r} 구분={gubun!r}")


def _warn_skipped_blocks(skipped):
    """읍면동 라벨이 있는데도 통째로 건너뛴 블록을 사유별로 집계해 경고. 정상이면 0건.

    블록이 통째로 사라지는 누락(경북 ~1.73M표 사례)이 다시는 조용히 일어나지 않도록,
    skip된 모든 블록을 선거종류/시도/구시군/source/사유와 함께 드러낸다."""
    counts = Counter(
        (block["reason"], block["선거종류"], block["시도"], block["구시군"])
        for block in skipped)
    print(f"[parse_6th] 경고: 블록 통째 건너뜀 {len(skipped)}건 "
          f"({len(counts)}종) — 데이터 누락 의심, 헤더 매칭 점검 필요:")
    for (reason, etype, sido, gu), count in counts.most_common(30):
        sample = next(block["source"] for block in skipped
                      if block["reason"] == reason and block["선거종류"] == etype
                      and block["시도"] == sido and block["구시군"] == gu)
        print(f"  {count:4d}블록  사유={reason} 선거종류={etype} 시도={sido} "
              f"구시군={gu} 예시={sample!r}")
