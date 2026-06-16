"""22대 국회의원선거 파서 (지역구 + 비례대표).

원본 파일:
  data_raw/개표결과/1. 개표단위별 개표결과(지역구) -전국.xlsx
  data_raw/개표결과/2. 개표단위별 개표결과(비례대표) -전국.xlsx

반환값: (rows: list[dict], totals: list[dict])
  rows  — COLUMNS 스키마를 따르는 tidy 행
  totals — 지역구 선거구별 합계행 (선거구분/시도/선거구명/투표수)
"""

import os
import openpyxl

from etl.assembly.schema import COLUMNS, normalize_level
from etl.assembly._helpers import to_int, normalize_sido, parse_party_candidate, is_skip_row

RAW_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data_raw")


def _parse_22nd_지역구(raw_dir):
    path = os.path.join(raw_dir, "개표결과", "1. 개표단위별 개표결과(지역구) -전국.xlsx")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        print(f"[SKIP] 22대 지역구: {error}")
        return [], []
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 5:
        return [], []

    header_row = list(all_rows[3])
    ncols = len(header_row)

    시도_col = 0
    선거구명_col = 1
    구시군_col = 2
    emd_col = 3
    투표타입_col = 4
    선거인수_col = 5
    투표수_col = 6
    cand_start = 7

    무효_col = None
    기권_col = None
    for col_idx in range(ncols - 1, -1, -1):
        header_val = str(header_row[col_idx] or "").replace("\n", "").strip()
        if "기권" in header_val and 기권_col is None:
            기권_col = col_idx
        if "무효" in header_val and 무효_col is None:
            무효_col = col_idx

    end_col = min(c for c in [무효_col, 기권_col] if c is not None) if (무효_col or 기권_col) else ncols

    rows = []
    # totals_raw: (시도, 선거구명, 구시군) → 투표수 누적 (구시군별 합산 후 선거구명 단위로 집계)
    totals_raw = {}

    current_sido = None
    current_선거구 = None
    current_구시군 = None
    current_candidates = []
    current_emd = ""

    for r_idx in range(4, len(all_rows)):
        row = all_rows[r_idx]
        if len(row) < 7:
            continue

        sido_val = str(row[시도_col] or "").strip()
        선거구_val = str(row[선거구명_col] or "").strip()
        구시군_val = str(row[구시군_col] or "").strip()
        emd_val = str(row[emd_col] or "").strip()
        타입_val = str(row[투표타입_col] or "").strip()

        # 선거구 정의행: 선거구명이 있고 선거인수가 비어있는 행 (후보자 정보 행)
        # 첫 선거구는 시도가 함께 표기되고, 이후 같은 시도 선거구는 선거구명만 표기됨
        if (sido_val or 선거구_val) and not row[선거인수_col]:
            if sido_val:
                current_sido = normalize_sido(sido_val)
            current_선거구 = 선거구_val
            current_구시군 = 구시군_val
            current_emd = ""
            current_candidates = []
            for col_idx in range(cand_start, end_col):
                raw = str(row[col_idx] or "").strip()
                if not raw or raw == "계":
                    continue
                party, cand = parse_party_candidate(raw)
                if party is None and cand is None:
                    continue
                current_candidates.append((col_idx, party or "", cand or ""))
            continue

        # 구시군 소구분 행: 구시군만 있고 읍면동/타입/선거인수가 없는 행
        # (다중 구시군 선거구에서 각 구시군의 시작을 표시)
        if 구시군_val and not emd_val and not 타입_val and not row[선거인수_col]:
            current_구시군 = 구시군_val
            current_emd = ""
            continue

        # 소계행에서 읍면동명 갱신 (데이터행에는 빈 값으로 오는 경우가 많음)
        if emd_val and 타입_val == "소계":
            current_emd = emd_val

        if not current_sido:
            continue

        # 합계행: emd='합계'이고 선거인수가 있는 행 → 구시군별 투표수 누적
        if emd_val == "합계" and row[선거인수_col] is not None:
            agg_key = (current_sido, current_선거구, current_구시군)
            if agg_key not in totals_raw:
                totals_raw[agg_key] = to_int(row[투표수_col]) or 0
            continue

        # 일반 집계행 스킵
        if is_skip_row(emd_val) and is_skip_row(타입_val):
            continue
        if emd_val == "합계":
            continue

        투표구명 = 타입_val or emd_val
        # 읍면동: 타입_val이 있는 행(투표구명이 타입_val)에서는 emd_val을 사용,
        # emd_val이 없으면 직전 소계행의 읍면동을 사용
        if 타입_val:
            emd_actual = emd_val if emd_val else current_emd
        else:
            emd_actual = ""

        if is_skip_row(투표구명):
            continue

        선거인수 = to_int(row[선거인수_col])
        투표수 = to_int(row[투표수_col])
        무효 = to_int(row[무효_col]) if 무효_col is not None and 무효_col < len(row) else None
        기권 = to_int(row[기권_col]) if 기권_col is not None and 기권_col < len(row) else None

        for col_idx, party, cand in current_candidates:
            득표 = to_int(row[col_idx]) if col_idx < len(row) else None
            rows.append({
                "선거_회차": "제22대",
                "선거일": "2024-04-10",
                "선거구분": "지역구",
                "시도": current_sido,
                "구시군": current_구시군,
                "읍면동": emd_actual,
                "투표구": 투표구명,
                "선거구명": current_선거구,
                "선거인수": 선거인수,
                "투표수": 투표수,
                "후보자": cand,
                "정당": party,
                "득표수": 득표,
                "무효투표수": 무효,
                "기권수": 기권,
                "level": normalize_level(투표구명),
            })

    # 구시군별 투표수를 (시도, 선거구명) 단위로 합산
    from collections import defaultdict
    aggregated = defaultdict(int)
    for (sido, 선거구명, 구시군), 투표수 in totals_raw.items():
        aggregated[(sido, 선거구명)] += 투표수

    totals = [
        {"선거구분": "지역구", "시도": sido, "선거구명": 선거구명, "투표수": 합계}
        for (sido, 선거구명), 합계 in aggregated.items()
    ]
    return rows, totals


def _parse_22nd_비례(raw_dir):
    path = os.path.join(raw_dir, "개표결과", "2. 개표단위별 개표결과(비례대표) -전국.xlsx")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        print(f"[SKIP] 22대 비례: {error}")
        return [], []
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 7:
        return [], []

    # row4 (idx=3): 컬럼헤더, row5 (idx=4): 정당명
    header_row = list(all_rows[3])
    party_row = list(all_rows[4])
    ncols = len(header_row)

    시도_col = 0
    구시군_col = 1
    emd_col = 2
    투표구_col = 3
    선거인수_col = 4
    투표수_col = 5
    cand_start = 6

    무효_col = None
    기권_col = None
    for col_idx in range(ncols - 1, -1, -1):
        header_val = str(header_row[col_idx] or "").replace("\n", "").strip()
        if "기권" in header_val and 기권_col is None:
            기권_col = col_idx
        if "무효" in header_val and 무효_col is None:
            무효_col = col_idx

    end_col = min(c for c in [무효_col, 기권_col] if c is not None) if (무효_col or 기권_col) else ncols

    candidate_cols = []
    for col_idx in range(cand_start, end_col):
        party = str(party_row[col_idx] or "").strip()
        if party in ("계", "", "None"):
            continue
        candidate_cols.append((col_idx, party))

    rows = []
    current_sido = None
    current_구시군 = None
    current_emd = ""

    for r_idx in range(6, len(all_rows)):
        row = all_rows[r_idx]
        if len(row) < 6:
            continue

        sido_val = str(row[시도_col] or "").strip()
        구시군_val = str(row[구시군_col] or "").strip()
        emd_val = str(row[emd_col] or "").strip()
        투표구명 = str(row[투표구_col] or "").strip()

        if sido_val:
            current_sido = normalize_sido(sido_val)
        if 구시군_val:
            current_구시군 = 구시군_val

        # 소계행에서 읍면동명 갱신
        if emd_val and 투표구명 in ("소계", ""):
            current_emd = emd_val

        if is_skip_row(emd_val) and is_skip_row(투표구명):
            continue
        if is_skip_row(투표구명) and not emd_val:
            continue
        if is_skip_row(투표구명):
            continue

        # 데이터행의 읍면동이 비어있으면 직전 소계행의 읍면동을 사용
        if not emd_val:
            emd_val = current_emd

        if not current_sido:
            continue

        선거인수 = to_int(row[선거인수_col])
        투표수 = to_int(row[투표수_col])
        무효 = to_int(row[무효_col]) if 무효_col is not None and 무효_col < len(row) else None
        기권 = to_int(row[기권_col]) if 기권_col is not None and 기권_col < len(row) else None

        for col_idx, party in candidate_cols:
            득표 = to_int(row[col_idx]) if col_idx < len(row) else None
            rows.append({
                "선거_회차": "제22대",
                "선거일": "2024-04-10",
                "선거구분": "비례대표",
                "시도": current_sido,
                "구시군": current_구시군,
                "읍면동": emd_val,
                "투표구": 투표구명,
                "선거구명": "",
                "선거인수": 선거인수,
                "투표수": 투표수,
                "후보자": "",
                "정당": party,
                "득표수": 득표,
                "무효투표수": 무효,
                "기권수": 기권,
                "level": normalize_level(투표구명),
            })

    return rows, []


def parse_22nd(raw_dir=RAW_DIR):
    """22대 지역구+비례 파싱.

    Returns:
        (rows, totals) — rows는 COLUMNS 스키마 dict 리스트,
                         totals는 지역구 선거구별 합계 dict 리스트.
    """
    rows_지역구, totals_지역구 = _parse_22nd_지역구(raw_dir)
    rows_비례, _ = _parse_22nd_비례(raw_dir)
    return rows_지역구 + rows_비례, totals_지역구
