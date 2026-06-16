"""
제22대 국회의원선거 개표결과 xlsx → CSV 변환
출력: data_processed/22대총선_지역구.csv
        data_processed/22대총선_비례대표.csv
        data_processed/22대총선_재보궐_구시군의장.csv
        data_processed/22대총선_재보궐_시도의회의원.csv
        data_processed/22대총선_재보궐_구시군의회의원.csv
"""

import csv
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).parents[2] / "data_raw" / "제22대_국회의원선거_개표결과"
OUT_DIR = Path(__file__).parents[2] / "data_processed"
OUT_DIR.mkdir(exist_ok=True)

FIXED_TAIL = {"계", "무효\n투표수", "기권수"}


def clean_num(value):
    if value is None:
        return ""
    return str(value).replace(",", "").strip()


def is_candidate(value):
    """'정당\n후보자' 형태인지 확인"""
    return value is not None and "\n" in str(value)


def parse_candidate(raw):
    parts = str(raw).split("\n", 1)
    return parts[0].strip(), parts[1].strip()


def convert_district(path: Path, out_path: Path):
    """
    지역구 / 재보궐 파일 변환.

    선거구 헤더 행: col0=시도명, col1=선거구명, col2=구시군명, col7~=후보자('정당\\n이름')
    데이터 행: col0~2=None, col3=읍면동명, col4=투표타입, col5=선거인수, col6=투표수, col7~=득표수
    마지막 고정 열: 계(col57), 무효투표수(col58), 기권수(col59)
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    sido = sgg = gusi = ""
    candidates = []   # [(col_idx, 정당, 후보자명)]
    tail = {}         # {"계": col_idx, "무효투표수": col_idx, "기권수": col_idx}

    for row in rows:
        # 선거구 헤더 행: col1에 선거구명, col7 이상에 후보자 컬럼 ('정당\n이름' 형태)
        if row[1] is not None and row[3] is None and len(row) > 7 and is_candidate(row[7]):
            sido = str(row[0]).strip() if row[0] else sido  # col0 없으면 이전 시도 유지
            sgg = str(row[1]).strip()
            gusi = str(row[2]).strip() if row[2] else ""
            candidates = []
            tail = {}
            for col_idx, val in enumerate(row):
                if col_idx < 7:
                    continue
                if val is None:
                    continue
                label = str(val).strip()
                if label in FIXED_TAIL or label.replace("\n", "") in {"계", "무효투표수", "기권수"}:
                    key = "무효투표수" if "무효" in label else label.replace("\n", "")
                    tail[key] = col_idx
                elif is_candidate(val):
                    party, name = parse_candidate(val)
                    candidates.append((col_idx, party, name))
            continue

        # 고정 꼬리 열 찾기 (헤더 다음 행에 '계' 등이 나오는 경우)
        if not tail:
            for col_idx, val in enumerate(row):
                if val in FIXED_TAIL:
                    key = "무효투표수" if "무효" in str(val) else str(val).replace("\n", "")
                    tail[key] = col_idx

        # 데이터 행: col3(읍면동) 또는 col4(투표타입)에 값이 있어야 함
        emd = row[3]
        vote_type = row[4]
        sunsu = clean_num(row[5])
        tusu = clean_num(row[6])

        if (emd is None and vote_type is None) or (sunsu == "" and tusu == ""):
            continue
        if not candidates:
            continue

        for col_idx, party, name in candidates:
            votes = clean_num(row[col_idx]) if col_idx < len(row) else ""
            records.append({
                "시도명": sido,
                "선거구명": sgg,
                "구시군명": gusi,
                "읍면동명": str(emd).strip() if emd else "",
                "투표타입": str(vote_type).strip() if vote_type else "",
                "선거인수": sunsu,
                "투표수": tusu,
                "정당": party,
                "후보자명": name,
                "득표수": votes,
                "계": clean_num(row[tail["계"]]) if "계" in tail and tail["계"] < len(row) else "",
                "무효투표수": clean_num(row[tail["무효투표수"]]) if "무효투표수" in tail and tail["무효투표수"] < len(row) else "",
                "기권수": clean_num(row[tail["기권수"]]) if "기권수" in tail and tail["기권수"] < len(row) else "",
            })

    fieldnames = ["시도명", "선거구명", "구시군명", "읍면동명", "투표타입",
                  "선거인수", "투표수", "정당", "후보자명", "득표수",
                  "계", "무효투표수", "기권수"]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    print(f"  저장: {out_path.name} ({len(records):,}행)")


def convert_proportional(path: Path, out_path: Path):
    """
    비례대표 파일 변환.

    row4: 정당명 헤더 (col6~, '계' 제외)
    row6~: 데이터 (col0=시도명, col1=구시군명, col2=읍면동명, col3=투표구명,
                   col4=선거인수, col5=투표수, col6~=득표수)
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    party_row = rows[4]
    parties = [
        (col_idx, str(val).strip())
        for col_idx, val in enumerate(party_row)
        if col_idx >= 6 and val is not None and str(val).strip() not in {"계", ""}
    ]

    records = []
    for row in rows[6:]:
        sunsu = clean_num(row[4])
        tusu = clean_num(row[5])
        if sunsu == "" and tusu == "":
            continue

        for col_idx, party in parties:
            votes = clean_num(row[col_idx]) if col_idx < len(row) else ""
            records.append({
                "시도명": str(row[0]).strip() if row[0] else "",
                "구시군명": str(row[1]).strip() if row[1] else "",
                "읍면동명": str(row[2]).strip() if row[2] else "",
                "투표구명": str(row[3]).strip() if row[3] else "",
                "선거인수": sunsu,
                "투표수": tusu,
                "정당": party,
                "득표수": votes,
            })

    fieldnames = ["시도명", "구시군명", "읍면동명", "투표구명", "선거인수", "투표수", "정당", "득표수"]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    print(f"  저장: {out_path.name} ({len(records):,}행)")


def main():
    print("제22대 국회의원선거 개표결과 변환 시작")

    convert_district(
        RAW_DIR / "1. 개표단위별 개표결과(지역구) -전국.xlsx",
        OUT_DIR / "22대총선_지역구.csv",
    )
    convert_proportional(
        RAW_DIR / "2. 개표단위별 개표결과(비례대표) -전국.xlsx",
        OUT_DIR / "22대총선_비례대표.csv",
    )
    convert_district(
        RAW_DIR / "3. 개표단위별 개표결과(재보궐) -구시군의장.xlsx",
        OUT_DIR / "22대총선_재보궐_구시군의장.csv",
    )
    convert_district(
        RAW_DIR / "4. 개표단위별 개표결과(재보궐) -시도의회의원.xlsx",
        OUT_DIR / "22대총선_재보궐_시도의회의원.csv",
    )
    convert_district(
        RAW_DIR / "5. 개표단위별 개표결과(재보궐) -구시군의회의원.xlsx",
        OUT_DIR / "22대총선_재보궐_구시군의회의원.csv",
    )

    print("완료")


if __name__ == "__main__":
    main()
