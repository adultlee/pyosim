"""
국회의원선거 raw 데이터 → 통합 CSV 정제 스크립트
출력: data_processed/국회의원선거.csv
"""
import os
import re
import xlrd
import openpyxl
import pandas as pd

BASE = "/Users/seong-in/Desktop/Git/pyosim"
OUT_CSV = os.path.join(BASE, "data_processed", "국회의원선거.csv")

FINAL_COLS = [
    "선거_회차", "선거일", "선거구분", "시도", "구시군", "읍면동", "투표구",
    "선거구명", "선거인수", "투표수", "후보자", "정당", "득표수", "무효투표수", "기권수", "level",
]

SKIP_ROW_KEYWORDS = {"합계", "소계", "부재자", "부재자투표", "잘못투입된투표지", "계", "잘못된투표지",
                     "국내부재자투표"}

LEVEL_MAP = {
    "관내사전투표": "사전투표",
    "관외사전투표": "관외사전투표",
    "거소·선상투표": "거소선상",
    "국외부재자투표": "재외투표",
    "재외투표": "재외투표",
}


def get_level(투표구명):
    """투표구명으로 level 판단. 일반 투표구는 '당일투표'."""
    s = str(투표구명).strip() if 투표구명 else ""
    return LEVEL_MAP.get(s, "당일투표")


def to_int(val):
    if val is None or val == "":
        return None
    try:
        cleaned = str(val).replace(",", "").strip()
        # Handle float strings like "2736.0"
        return int(float(cleaned))
    except (ValueError, TypeError):
        return None


def parse_party_candidate(col_str):
    """'정당\n후보자' 또는 '정당' 형식 → (정당, 후보자)
    newline이 있으면 앞=정당, 뒤=후보자
    newline이 없으면 전체=정당 (비례 케이스)
    """
    if not col_str or str(col_str).strip() in ("", "\n", "계"):
        return None, None
    raw = str(col_str)
    if "\n" in raw:
        parts = [p.strip() for p in raw.split("\n")]
        parts = [p for p in parts if p]
        if len(parts) == 2:
            return parts[0], parts[1]
        if len(parts) == 1:
            return parts[0], ""
        return None, None
    else:
        # single string → treat as party name (비례 케이스)
        return raw.strip(), ""


def is_skip_row(투표구명):
    """합계/소계 등 집계행 여부 판단"""
    if not 투표구명:
        return True
    s = str(투표구명).strip()
    if not s:
        return True
    if s in SKIP_ROW_KEYWORDS:
        return True
    # 소계 포함 패턴
    if "소계" in s or "합계" in s:
        return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# 16대 파서
# 구조: 단일 xls, 섹션마다 '#선거구명' 헤더 행 → 3행 헤더(정당행1, 정당행2, 후보자명)
# ─────────────────────────────────────────────────────────────────────────────

# 시도명 추정용 선거구→시도 매핑 (16대 전용)
SIDO_16 = {
    "서울": ["종로구", "중구", "용산구", "성동구", "광진구", "동대문구", "중랑구", "성북구",
             "강북구", "도봉구", "노원구", "은평구", "서대문구", "마포구", "양천구", "강서구",
             "구로구", "금천구", "영등포구", "동작구", "관악구", "서초구", "강남구", "송파구", "강동구"],
    "부산": ["중구(부산)", "서구(부산)", "동구", "영도구", "부산진구", "동래구", "남구", "북구강서구",
             "해운대구기장군", "사하구", "금정구", "연제구", "수영구", "사상구"],
    "대구": ["중구(대구)", "동구(대구)", "서구(대구)", "남구(대구)", "북구(대구)", "수성구", "달서구", "달성군"],
    "인천": ["중구(인천)", "동구(인천)", "남구(인천)", "연수구", "남동구", "부평구", "계양구", "서구(인천)", "강화군옹진군"],
    "광주": ["동구(광주)", "서구(광주)", "남구(광주)", "북구(광주)", "광산구"],
    "대전": ["동구(대전)", "중구(대전)", "서구(대전)", "유성구", "대덕구"],
    "울산": ["중구(울산)", "남구(울산)", "동구(울산)", "북구(울산)", "울주군"],
    "경기": ["수원시", "성남시", "의정부시", "안양시", "부천시", "광명시", "평택시", "동두천시",
             "안산시", "고양시", "과천시", "구리시", "남양주시", "오산시", "시흥시", "군포시",
             "의왕시", "하남시", "용인시", "파주시", "이천시", "안성시", "김포시", "화성시",
             "광주시", "양주시", "여주군", "연천군", "포천군", "가평군", "양평군"],
    "강원": ["춘천시", "원주시", "강릉시", "동해시", "태백시", "속초시", "삼척시", "홍천군",
             "횡성군", "영월군", "평창군", "정선군", "철원군", "화천군", "양구군", "인제군", "고성군양양군"],
    "충북": ["청주시", "충주시", "제천시", "청원군", "보은군", "옥천군", "영동군", "진천군",
             "괴산군증평군", "음성군", "단양군"],
    "충남": ["천안시", "공주시", "보령시", "아산시", "서산시", "논산시", "연기군", "금산군",
             "연기군", "부여군", "서천군", "청양군", "홍성군", "예산군", "태안군", "당진군"],
    "전북": ["전주시", "군산시", "익산시", "정읍시", "남원시", "김제시", "완주군", "진안군",
             "무주군", "장수군", "임실군", "순창군", "고창군", "부안군"],
    "전남": ["목포시", "여수시", "순천시", "나주시", "광양시구례군", "담양군곡성군장성군",
             "고흥군보성군", "화순군", "장흥군강진군", "해남군완도군진도군", "무안군신안군", "영암군",
             "함평군영광군"],
    "경북": ["포항시", "경주시", "김천시", "안동시", "구미시", "영주시", "영천시", "상주시",
             "문경시", "경산시", "군위군의성군청송군", "영양군울진군", "영덕군", "청도군고령군",
             "성주군칠곡군", "예천군", "봉화군", "울릉군독도"],
    "경남": ["창원시", "마산시", "진주시", "진해시", "통영시고성군", "사천시", "김해시", "밀양시",
             "거제시", "의령군함안군", "창녕군", "고성군", "하동군", "산청군합천군", "남해군하동군",
             "함양군거창군", "양산시"],
    "제주": ["제주시", "북제주군", "서귀포시,남제주군"],
}


SIDO_16_EXTRA = {
    "중동구": "부산",
    "남원·순창": "전북", "진안·무주·장수": "전북", "고창·부안": "전북", "완주·임실": "전북",
    "고흥군": "전남", "장흥군영암군": "전남", "해남군진도군": "전남",
    "보성군화순군": "전남", "강진군완도군": "전남",
    "고령군성주군": "경북", "군위군의성군": "경북", "청송군영양군영덕군": "경북", "칠곡군": "경북",
}


def infer_sido_16(선거구명_raw):
    """16대용: 선거구명 raw string에서 시도 추정"""
    name = str(선거구명_raw).strip()
    # 직접 매핑 먼저
    if name in SIDO_16_EXTRA:
        return SIDO_16_EXTRA[name]
    for sido, prefixes in SIDO_16.items():
        for prefix in prefixes:
            if name.startswith(prefix.split("(")[0]):
                return sido
    # 이름 자체에 시도 포함 단서 존재
    if "서울" in name:
        return "서울"
    if "부산" in name:
        return "부산"
    if "대구" in name:
        return "대구"
    if "인천" in name:
        return "인천"
    if "광주" in name:
        return "광주"
    if "대전" in name:
        return "대전"
    if "울산" in name:
        return "울산"
    return ""


def parse_16대():
    path = os.path.join(BASE, "data_raw",
                        "국회의원선거 개표결과(제16대~19대)",
                        "제16대 국회의원선거",
                        "제16대국회의원선거투표구별득표상황.xls")
    wb = xlrd.open_workbook(path, encoding_override="cp949")
    ws = wb.sheet_by_index(0)
    rows = []

    # 구간 수집: '#선거구명' 으로 구분
    section_starts = []
    for i in range(ws.nrows):
        val = str(ws.cell_value(i, 0))
        if val.startswith("#"):
            section_starts.append((i, val[1:].strip()))

    for sec_idx, (sec_row, 선거구명_raw) in enumerate(section_starts):
        # 다음 섹션 시작 전까지
        next_row = section_starts[sec_idx + 1][0] if sec_idx + 1 < len(section_starts) else ws.nrows
        sido = infer_sido_16(선거구명_raw)
        선거구명 = 선거구명_raw

        # 헤더 파싱: sec_row+1 = 컬럼행(투표구명/선거인수/투표수/.../#후별득.../무효/기권수)
        # sec_row+2,+3,+4 = 정당행1, 정당행2, 후보자행
        header_row = sec_row + 1
        party_row1 = sec_row + 2
        party_row2 = sec_row + 3
        candidate_row = sec_row + 4

        if party_row2 >= ws.nrows or candidate_row >= ws.nrows:
            continue

        # 컬럼 구조: col0=투표구명, col1=선거인수, col2=투표수, col3..n-3=후보, n-2=계, n-1=무효, n=기권수
        ncols = ws.ncols
        # 무효투표: col n-2 (0-indexed: ncols-2), 기권: col ncols-1
        # 계: col ncols-3
        # 후보 컬럼: 3 ~ ncols-4

        # 후보 정당명 = party_row1 + party_row2 (두 행 결합)
        # 후보자명 = candidate_row
        candidate_cols = []
        for c in range(3, ncols - 3):
            p1 = str(ws.cell_value(party_row1, c)).strip()
            p2 = str(ws.cell_value(party_row2, c)).strip()
            cand = str(ws.cell_value(candidate_row, c)).strip()
            party = (p1 + " " + p2).strip()
            # 공백 정규화
            party = re.sub(r"\s+", " ", party).strip()
            if party or cand:
                candidate_cols.append((c, party, cand))

        data_start = candidate_row + 1
        for r in range(data_start, next_row):
            투표구명 = str(ws.cell_value(r, 0)).strip()
            if is_skip_row(투표구명):
                continue
            if 투표구명.startswith("#") or 투표구명 in ("투표구명", ""):
                continue

            선거인수 = to_int(ws.cell_value(r, 1))
            투표수 = to_int(ws.cell_value(r, 2))
            무효 = to_int(ws.cell_value(r, ncols - 2))
            기권 = to_int(ws.cell_value(r, ncols - 1))

            for c, party, cand in candidate_cols:
                득표 = to_int(ws.cell_value(r, c))
                if not cand and not party:
                    continue
                rows.append({
                    "선거_회차": "제16대",
                    "선거일": "2000-04-13",
                    "선거구분": "지역구",
                    "시도": sido,
                    "구시군": "",
                    "읍면동": "",
                    "투표구": 투표구명,
                    "선거구명": 선거구명,
                    "선거인수": 선거인수,
                    "투표수": 투표수,
                    "후보자": cand,
                    "정당": party,
                    "득표수": 득표,
                    "무효투표수": 무효,
                    "기권수": 기권,
                    "level": get_level(투표구명),
                })

    return pd.DataFrame(rows, columns=FINAL_COLS)


# ─────────────────────────────────────────────────────────────────────────────
# 17대 파서 (지역구 + 비례대표)
# 구조: 시도별 xls, 각 xls에 여러 시트 (시트명=선거구명)
# 헤더: row0=컬럼명, row1=정당, row2=후보자 (지역구) 또는 row0=컬럼, row1=정당 (비례)
# ─────────────────────────────────────────────────────────────────────────────

SIDO_MAP_17 = {
    "01 서울": "서울", "02 부산": "부산", "03 대구": "대구", "04 인천": "인천",
    "05 광주": "광주", "06 대전": "대전", "07 울산": "울산", "08 경기": "경기",
    "09 강원": "강원", "10 충북": "충북", "11 충남": "충남", "12 전북": "전북",
    "13 전남": "전남", "14 경북": "경북", "15 경남": "경남", "16 제주": "제주",
}


def parse_17대_시도_file(path, sido, 선거구분, 선거일="2004-04-15", 회차="제17대"):
    """17대/18대 지역구 공통: 시도별 xls, 시트별로 선거구"""
    wb = xlrd.open_workbook(path, encoding_override="cp949")
    rows = []
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        선거구명 = ws.name.strip()
        ncols = ws.ncols
        if ws.nrows < 4:
            continue

        # row0: 컬럼헤더 (투표구명, 선거인수, 투표수, ...)
        # row1: 정당
        # row2: 후보자명 (지역구만)
        # 비례는 row1이 정당, row2부터 데이터

        # 마지막 컬럼들 파악: '무효투표수', '기권수' 위치
        header_row0 = [str(ws.cell_value(0, c)).strip() for c in range(ncols)]
        header_row1 = [str(ws.cell_value(1, c)).strip() for c in range(ncols)]
        header_row2 = [str(ws.cell_value(2, c)).strip() for c in range(ncols)]

        # 컬럼 오프셋 파악
        # 지역구: 0=투표구명, 1=선거인수, 2=투표수
        # 비례: 0=투표구명, 1=선거인수, 2=투표수 (동일)
        # 17대 지역구는 row0,1,2 이후 row3부터 데이터
        # 18대 지역구는 읍면동명이 col0으로 추가됨

        # col 오프셋 감지
        if header_row0[0] in ("읍면동명",):
            # 18대 지역구 형식: col0=읍면동, col1=투표구명, col2=선거인수, col3=투표수
            emd_col = 0
            투표구_col = 1
            선거인수_col = 2
            투표수_col = 3
            cand_start_col = 4
            data_start_row = 3
        else:
            # 17대 형식: col0=투표구명, col1=선거인수, col2=투표수
            emd_col = None
            투표구_col = 0
            선거인수_col = 1
            투표수_col = 2
            cand_start_col = 3
            data_start_row = 3

        # 무효/기권 컬럼 찾기
        무효_col = None
        기권_col = None
        for c in range(ncols - 1, -1, -1):
            v0 = str(ws.cell_value(0, c)).strip()
            if "기권" in v0 and 기권_col is None:
                기권_col = c
            if "무효" in v0 and 무효_col is None:
                무효_col = c

        # 후보 컬럼 파악: cand_start_col ~ (무효_col or 기권_col or ncols) - 2
        end_col = min(c for c in [무효_col, 기권_col] if c is not None) if (무효_col or 기권_col) else ncols
        # 계 컬럼 제외
        candidate_cols = []
        for c in range(cand_start_col, end_col):
            if 선거구분 == "지역구":
                party = re.sub(r"\s+", " ", str(ws.cell_value(1, c))).strip()
                cand = re.sub(r"\s+", " ", str(ws.cell_value(2, c))).strip()
                combined = f"{party}\n{cand}" if party else cand
            else:
                # 비례: row1=정당
                party = re.sub(r"\s+", " ", str(ws.cell_value(1, c))).strip()
                cand = party
                combined = party

            if not party and not cand:
                continue
            if party in ("", "계") and cand in ("", "계"):
                continue
            candidate_cols.append((c, party, cand))

        for r in range(data_start_row, ws.nrows):
            투표구명 = str(ws.cell_value(r, 투표구_col)).strip()
            if is_skip_row(투표구명):
                continue
            emd = str(ws.cell_value(r, emd_col)).strip() if emd_col is not None else ""
            선거인수 = to_int(ws.cell_value(r, 선거인수_col))
            투표수 = to_int(ws.cell_value(r, 투표수_col))
            무효 = to_int(ws.cell_value(r, 무효_col)) if 무효_col is not None else None
            기권 = to_int(ws.cell_value(r, 기권_col)) if 기권_col is not None else None

            for c, party, cand in candidate_cols:
                득표 = to_int(ws.cell_value(r, c))
                actual_cand = cand if 선거구분 == "지역구" else ""
                actual_party = party
                rows.append({
                    "선거_회차": 회차,
                    "선거일": 선거일,
                    "선거구분": 선거구분,
                    "시도": sido,
                    "구시군": "",
                    "읍면동": emd,
                    "투표구": 투표구명,
                    "선거구명": 선거구명,
                    "선거인수": 선거인수,
                    "투표수": 투표수,
                    "후보자": actual_cand,
                    "정당": actual_party,
                    "득표수": 득표,
                    "무효투표수": 무효,
                    "기권수": 기권,
                    "level": get_level(투표구명),
                })

    return pd.DataFrame(rows, columns=FINAL_COLS)


def parse_17대():
    base17 = os.path.join(BASE, "data_raw", "국회의원선거 개표결과(제16대~19대)", "제17대 국회의원선거")
    dfs = []
    for fname_stem, sido in SIDO_MAP_17.items():
        for 구분 in ["지역구", "비례대표"]:
            path = os.path.join(base17, f"제17대 국회의원선거({구분})", f"{fname_stem}.xls")
            if os.path.exists(path):
                df = parse_17대_시도_file(path, sido, 구분, "2004-04-15", "제17대")
                dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(columns=FINAL_COLS)


# ─────────────────────────────────────────────────────────────────────────────
# 18대 파서
# 지역구: 시도별 xls, 시트별 선거구 (읍면동명 컬럼 있음)
# 비례대표: 단일 xls, 시트별 시도, 구조: col0=구시군명, col1=읍면동명, col2=투표구명
# ─────────────────────────────────────────────────────────────────────────────

SIDO_MAP_18 = {
    "서울": "서울", "부산": "부산", "대구": "대구", "인천": "인천", "광주": "광주",
    "대전": "대전", "울산": "울산", "경기": "경기", "강원": "강원", "충북": "충북",
    "충남": "충남", "전북": "전북", "전남": "전남", "경북": "경북", "경남": "경남", "제주": "제주",
}


def parse_18대_비례():
    path = os.path.join(BASE, "data_raw", "국회의원선거 개표결과(제16대~19대)", "제18대 국회의원선거",
                        "제18대 국회의원선거(비례대표).xls")
    wb = xlrd.open_workbook(path, encoding_override="cp949")
    rows = []

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        sido = ws.name.strip()
        ncols = ws.ncols
        if ws.nrows < 3:
            continue

        # row0: col0=구시군명, col1=읍면동명, col2=투표구명, col3=선거인수, col4=투표수, col5..=정당
        # row1: 정당명들
        header0 = [str(ws.cell_value(0, c)).strip() for c in range(ncols)]
        party_row1 = [str(ws.cell_value(1, c)).strip() for c in range(ncols)]

        구시군_col = 0
        emd_col = 1
        투표구_col = 2
        선거인수_col = 3
        투표수_col = 4
        cand_start = 5

        무효_col = None
        기권_col = None
        for c in range(ncols - 1, -1, -1):
            v = str(ws.cell_value(0, c)).strip()
            if "기권" in v and 기권_col is None:
                기권_col = c
            if "무효" in v and 무효_col is None:
                무효_col = c

        end_col = min(c for c in [무효_col, 기권_col] if c is not None) if (무효_col or 기권_col) else ncols
        candidate_cols = []
        for c in range(cand_start, end_col):
            party = party_row1[c]
            if party in ("계", "", "무효\n투표수"):
                continue
            candidate_cols.append((c, party))

        for r in range(2, ws.nrows):
            구시군 = str(ws.cell_value(r, 구시군_col)).strip()
            emd = str(ws.cell_value(r, emd_col)).strip()
            투표구명 = str(ws.cell_value(r, 투표구_col)).strip()

            # 합계/소계 행 제거: 투표구명이 없거나 집계 키워드면 스킵
            if is_skip_row(투표구명):
                continue
            if not 투표구명:
                continue

            선거인수 = to_int(ws.cell_value(r, 선거인수_col))
            투표수 = to_int(ws.cell_value(r, 투표수_col))
            무효 = to_int(ws.cell_value(r, 무효_col)) if 무효_col is not None else None
            기권 = to_int(ws.cell_value(r, 기권_col)) if 기권_col is not None else None

            for c, party in candidate_cols:
                득표 = to_int(ws.cell_value(r, c))
                rows.append({
                    "선거_회차": "제18대",
                    "선거일": "2008-04-09",
                    "선거구분": "비례대표",
                    "시도": sido,
                    "구시군": 구시군,
                    "읍면동": emd,
                    "투표구": 투표구명,
                    "선거구명": "",
                    "선거인수": 선거인수,
                    "투표수": 투표수,
                    "후보자": "",
                    "정당": party,
                    "득표수": 득표,
                    "무효투표수": 무효,
                    "기권수": 기권,
                    "level": get_level(투표구명),
                })

    return pd.DataFrame(rows, columns=FINAL_COLS)


def parse_18대():
    base18 = os.path.join(BASE, "data_raw", "국회의원선거 개표결과(제16대~19대)", "제18대 국회의원선거")
    dfs = []
    # 지역구
    for fname, sido in SIDO_MAP_18.items():
        path = os.path.join(base18, "제18대 국회의원선거(지역구)", f"{fname}.xls")
        if os.path.exists(path):
            df = parse_17대_시도_file(path, sido, "지역구", "2008-04-09", "제18대")
            dfs.append(df)
    # 비례대표
    dfs.append(parse_18대_비례())
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(columns=FINAL_COLS)


# ─────────────────────────────────────────────────────────────────────────────
# 19대 파서
# 지역구/비례: 구별 파일 (xls), 구조: 몇 행 제목 후 row3=헤더, row4=후보/정당
# ─────────────────────────────────────────────────────────────────────────────

def parse_19_20_21대_xls_file(path, 회차, 선거일, 선거구분):
    """
    19대 파일 구조 (xls):
      row0: '' '개표진행상황(투표구별)' ...
      row1: 빈 행
      row2: '' '[국회의원선거]' or '[비례대표국회의원선거]'
      row3: '' '읍면동명' '투표구명' '선거인수' '투표수' ... '무효투표수' '기권수'
      row4: '' '' '' '' '' '정당\n후보자' ...  (후보컬럼)
      row5~: 데이터
    """
    wb = xlrd.open_workbook(path, encoding_override="cp949")
    ws = wb.sheet_by_index(0)
    ncols = ws.ncols
    rows = []

    if ws.nrows < 6:
        return pd.DataFrame(columns=FINAL_COLS)

    # 파일명/행 3에서 선거구명, 시도, 구시군 추출
    title_row = str(ws.cell_value(2, 1)).strip() if ncols > 1 else ""
    # e.g. '[국회의원선거]' or '[비례대표국회의원선거]'

    header_row = 3
    cand_row = 4

    # 컬럼 찾기
    header = [str(ws.cell_value(header_row, c)).strip() for c in range(ncols)]
    # col1=읍면동명, col2=투표구명, col3=선거인수, col4=투표수
    emd_col = 1
    투표구_col = 2
    선거인수_col = 3
    투표수_col = 4
    cand_start = 5

    무효_col = None
    기권_col = None
    for c in range(ncols - 1, -1, -1):
        v = header[c].replace("\n", "").strip()
        if "기권" in v and 기권_col is None:
            기권_col = c
        if "무효" in v and 무효_col is None:
            무효_col = c

    end_col = min(c for c in [무효_col, 기권_col] if c is not None) if (무효_col or 기권_col) else ncols

    candidate_cols = []
    for c in range(cand_start, end_col):
        raw = str(ws.cell_value(cand_row, c)).strip()
        if not raw or raw == "계" or raw == "\n":
            continue
        party, cand = parse_party_candidate(raw)
        if party is None and cand is None:
            continue
        candidate_cols.append((c, party or "", cand or ""))

    # 파일명으로부터 시도/구시군/선거구명 추출
    fname_base = os.path.basename(path)
    # 19대 패턴: '비례_서울_강남구.xls', '국회의원_경북_고령군성주군칠곡군_성주군.xls'
    # 20대/21대: title row에 '[국회의원선거][시도][구시군][선거구명]' 형식
    sido_str = ""
    구시군_str = ""
    선거구명_str = ""

    # 먼저 title row에서 bracket 파싱 시도 (20대/21대 xls 형식)
    title3 = str(ws.cell_value(2, 1)).strip() if ncols > 1 else ""
    bracket_parts = re.findall(r"\[([^\]]+)\]", title3)
    if len(bracket_parts) >= 3:
        # 형식: [선거종류][시도][구시군] or [선거종류][시도][구시군][선거구명]
        sido_str = normalize_sido(bracket_parts[1])
        구시군_str = bracket_parts[2]
        선거구명_str = bracket_parts[3] if len(bracket_parts) >= 4 else 구시군_str
    else:
        # 19대 형식: 파일명에서 추출
        # 예: '비례_서울_강남구.xls' → ['비례', '서울', '강남구']
        # 예: '국회의원_경북_고령군성주군칠곡군_성주군.xls' → [..., '경북', '고령군성주군칠곡군', '성주군']
        name_parts = fname_base.replace(".xls", "").replace(".xlsx", "").split("_")
        # 첫 부분은 '비례', '국회의원', '지역구' 등 선거종류
        # 두 번째가 시도, 세 번째가 구시군, 네 번째가 선거구명(있으면)
        if len(name_parts) >= 3:
            sido_str = normalize_sido(name_parts[1])
            구시군_str = name_parts[2]
            선거구명_str = name_parts[3] if len(name_parts) >= 4 else name_parts[2]
        elif len(name_parts) == 2:
            sido_str = normalize_sido(name_parts[1])

    data_start = 5
    for r in range(data_start, ws.nrows):
        emd = str(ws.cell_value(r, emd_col)).strip()
        투표구명 = str(ws.cell_value(r, 투표구_col)).strip()

        if is_skip_row(emd) and is_skip_row(투표구명):
            continue
        if is_skip_row(투표구명) and not emd:
            continue
        # 합계/소계 제거
        if is_skip_row(투표구명):
            continue

        선거인수 = to_int(ws.cell_value(r, 선거인수_col))
        투표수 = to_int(ws.cell_value(r, 투표수_col))
        무효 = to_int(ws.cell_value(r, 무효_col)) if 무효_col is not None else None
        기권 = to_int(ws.cell_value(r, 기권_col)) if 기권_col is not None else None

        for c, party, cand in candidate_cols:
            득표 = to_int(ws.cell_value(r, c))
            rows.append({
                "선거_회차": 회차,
                "선거일": 선거일,
                "선거구분": 선거구분,
                "시도": sido_str,
                "구시군": 구시군_str,
                "읍면동": emd,
                "투표구": 투표구명,
                "선거구명": 선거구명_str,
                "선거인수": 선거인수,
                "투표수": 투표수,
                "후보자": cand if 선거구분 == "지역구" else "",
                "정당": party,
                "득표수": 득표,
                "무효투표수": 무효,
                "기권수": 기권,
                "level": get_level(투표구명),
            })

    return pd.DataFrame(rows, columns=FINAL_COLS)


SIDO_NORMALIZE = {
    "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구",
    "인천광역시": "인천", "광주광역시": "광주", "대전광역시": "대전",
    "울산광역시": "울산", "세종특별자치시": "세종", "경기도": "경기",
    "강원도": "강원", "충청북도": "충북", "충청남도": "충남",
    "전라북도": "전북", "전라남도": "전남", "경상북도": "경북",
    "경상남도": "경남", "제주특별자치도": "제주",
    "서울": "서울", "부산": "부산", "대구": "대구", "인천": "인천",
    "광주": "광주", "대전": "대전", "울산": "울산", "세종": "세종",
    "경기": "경기", "강원": "강원", "충북": "충북", "충남": "충남",
    "전북": "전북", "전남": "전남", "경북": "경북", "경남": "경남", "제주": "제주",
    "강원특별자치도": "강원", "전북특별자치도": "전북",
}


def normalize_sido(raw):
    return SIDO_NORMALIZE.get(raw.strip(), raw.strip())


def parse_19대():
    base19 = os.path.join(BASE, "data_raw", "국회의원선거 개표결과(제16대~19대)", "제19대 국회의원선거")
    dfs = []
    for 구분 in ["지역구", "비례대표"]:
        dir_path = os.path.join(base19, f"제19대 국회의원선거({구분})")
        for root, dirs, files in os.walk(dir_path):
            for fname in sorted(files):
                if fname.endswith(".xls") or fname.endswith(".xlsx"):
                    path = os.path.join(root, fname)
                    df = parse_19_20_21대_xls_file(path, "제19대", "2012-04-11", 구분)
                    dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(columns=FINAL_COLS)


# ─────────────────────────────────────────────────────────────────────────────
# 20대/21대 파서 (xlsx, 유사한 구조)
# ─────────────────────────────────────────────────────────────────────────────

def parse_20_21대_xlsx_file(path, 회차, 선거일, 선거구분):
    """
    20대/21대 xlsx 구조:
      row1: '개표상황(투표구별)'
      row2: 출력일시
      row3: '[국회의원선거][시도명][구시군명][선거구명]'
      row4: '읍면동명' '투표구명' '선거인수' '투표수' ... '무효투표수' '기권수'
      row5: None None None None '정당\n후보자' ...  (후보컬럼)
      row6: 빈행 (있는 경우)
      row7~: 데이터
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [SKIP] {path}: {e}")
        return pd.DataFrame(columns=FINAL_COLS)
    ws = wb.active
    rows = []
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 6:
        return pd.DataFrame(columns=FINAL_COLS)

    # row3 (idx=2): title
    title3 = str(all_rows[2][0] or "").strip()
    bracket_parts = re.findall(r"\[([^\]]+)\]", title3)
    sido_str = ""
    구시군_str = ""
    선거구명_str = ""
    if len(bracket_parts) >= 2:
        sido_str = normalize_sido(bracket_parts[1])
        if len(bracket_parts) >= 3:
            구시군_str = bracket_parts[2]
        if len(bracket_parts) >= 4:
            선거구명_str = bracket_parts[3]
        else:
            선거구명_str = 구시군_str

    # row4 (idx=3): 컬럼헤더
    header_row = list(all_rows[3])
    ncols = len(header_row)

    emd_col = 0
    투표구_col = 1
    선거인수_col = 2
    투표수_col = 3
    cand_start = 4

    무효_col = None
    기권_col = None
    for c in range(ncols - 1, -1, -1):
        v = str(header_row[c] or "").replace("\n", "").strip()
        if "기권" in v and 기권_col is None:
            기권_col = c
        if "무효" in v and 무효_col is None:
            무효_col = c

    end_col = min(c for c in [무효_col, 기권_col] if c is not None) if (무효_col or 기권_col) else ncols

    # row5 (idx=4): 후보 컬럼
    cand_row = list(all_rows[4])
    candidate_cols = []
    for c in range(cand_start, end_col):
        raw = str(cand_row[c] or "").strip()
        if not raw or raw == "계" or raw == "\n":
            continue
        party, cand = parse_party_candidate(raw)
        if party is None and cand is None:
            continue
        candidate_cols.append((c, party or "", cand or ""))

    # 데이터 시작: row6(빈행) 혹은 row7부터
    data_start = 6
    # row6가 빈행인지 확인
    if len(all_rows) > 5 and all(v is None or str(v).strip() == "" for v in all_rows[5]):
        data_start = 6
    else:
        data_start = 5

    for r_idx in range(data_start, len(all_rows)):
        row = all_rows[r_idx]
        if len(row) < 4:
            continue
        emd = str(row[emd_col] or "").strip()
        투표구명 = str(row[투표구_col] or "").strip()

        if is_skip_row(emd) and is_skip_row(투표구명):
            continue
        if is_skip_row(투표구명):
            continue

        선거인수 = to_int(row[선거인수_col])
        투표수 = to_int(row[투표수_col])
        무효 = to_int(row[무효_col]) if 무효_col is not None and 무효_col < len(row) else None
        기권 = to_int(row[기권_col]) if 기권_col is not None and 기권_col < len(row) else None

        for c, party, cand in candidate_cols:
            득표 = to_int(row[c]) if c < len(row) else None
            rows.append({
                "선거_회차": 회차,
                "선거일": 선거일,
                "선거구분": 선거구분,
                "시도": sido_str,
                "구시군": 구시군_str,
                "읍면동": emd,
                "투표구": 투표구명,
                "선거구명": 선거구명_str,
                "선거인수": 선거인수,
                "투표수": 투표수,
                "후보자": cand if 선거구분 == "지역구" else "",
                "정당": party,
                "득표수": 득표,
                "무효투표수": 무효,
                "기권수": 기권,
                "level": get_level(투표구명),
            })

    return pd.DataFrame(rows, columns=FINAL_COLS)


def parse_20대():
    base20 = os.path.join(BASE, "data_raw", "제20대 국회의원선거 투표구별 개표결과")
    dfs = []
    for 구분 in ["지역구", "비례대표"]:
        dir_path = os.path.join(base20, 구분)
        for root, dirs, files in os.walk(dir_path):
            for fname in sorted(files):
                if fname.endswith(".xlsx"):
                    path = os.path.join(root, fname)
                    df = parse_20_21대_xlsx_file(path, "제20대", "2016-04-13", 구분)
                    dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(columns=FINAL_COLS)


def parse_21대():
    base21 = os.path.join(BASE, "data_raw", "제21대 국회의원선거(재보궐 포함) 투표구별 개표결과")
    dfs = []
    for 구분 in ["지역구", "비례대표"]:
        dir_path = os.path.join(base21, 구분)
        for root, dirs, files in os.walk(dir_path):
            for fname in sorted(files):
                if fname.endswith(".xlsx"):
                    path = os.path.join(root, fname)
                    df = parse_20_21대_xlsx_file(path, "제21대", "2020-04-15", 구분)
                    dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(columns=FINAL_COLS)


# ─────────────────────────────────────────────────────────────────────────────
# 22대 파서 (xlsx, 전국 단일 파일)
# ─────────────────────────────────────────────────────────────────────────────

def parse_22대_지역구():
    path = os.path.join(BASE, "data_raw", "개표결과", "1. 개표단위별 개표결과(지역구) -전국.xlsx")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[SKIP] 22대 지역구: {e}")
        return pd.DataFrame(columns=FINAL_COLS)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    rows = []

    if len(all_rows) < 5:
        return pd.DataFrame(columns=FINAL_COLS)

    # row4 (idx=3): 컬럼헤더
    # '시도명', '선거구명', '구시군명', '읍면동명', '투표타입', '선거인수', '투표수', '후보자별 득표수', ...
    header_row = list(all_rows[3])
    ncols = len(header_row)

    시도_col = 0
    선거구명_col = 1
    구시군_col = 2
    emd_col = 3
    투표타입_col = 4
    선거인수_col = 5
    투표수_col = 6
    cand_start = 7

    무효_col = None
    기권_col = None
    for c in range(ncols - 1, -1, -1):
        v = str(header_row[c] or "").replace("\n", "").strip()
        if "기권" in v and 기권_col is None:
            기권_col = c
        if "무효" in v and 무효_col is None:
            무효_col = c

    end_col = min(c for c in [무효_col, 기권_col] if c is not None) if (무효_col or 기권_col) else ncols

    # row5 (idx=4): 첫 선거구명 행 + 후보 컬럼
    # 22대: 선거구별 후보가 row5에 있고, 데이터는 row6부터
    # 하지만 여러 선거구가 있으므로 선거구별로 후보 컬럼이 다름
    # → 반복하면서 후보행(읍면동=None, 투표타입=None, 선거인수=None인데 후보컬럼에 값)을 감지

    current_sido = None
    current_선거구 = None
    current_구시군 = None
    current_candidates = []  # [(col_idx, party, cand), ...]

    for r_idx in range(4, len(all_rows)):
        row = all_rows[r_idx]
        if len(row) < 7:
            continue

        sido_val = str(row[시도_col] or "").strip()
        선거구_val = str(row[선거구명_col] or "").strip()
        구시군_val = str(row[구시군_col] or "").strip()
        emd_val = str(row[emd_col] or "").strip()
        타입_val = str(row[투표타입_col] or "").strip()

        # 후보행 감지: 시도/선거구에 값, 읍면동에 후보 또는 None 투표수
        # 22대 형식: row에 시도+선거구가 있고, 뒤에 후보 정보가 있는 행
        # 읍면동=None, 투표타입=None, 선거인수=None일 때가 선거구 정의행이 아니라 후보열
        if sido_val and 선거구_val:
            current_sido = normalize_sido(sido_val)
            current_선거구 = 선거구_val
            current_구시군 = 구시군_val
            # 이 행 자체가 후보 컬럼 정의행
            current_candidates = []
            for c in range(cand_start, end_col):
                raw = str(row[c] or "").strip()
                if not raw or raw == "계":
                    continue
                party, cand = parse_party_candidate(raw)
                if party is None and cand is None:
                    continue
                current_candidates.append((c, party or "", cand or ""))
            continue

        # 데이터행
        if not current_sido:
            continue

        # 투표구명 결정: emd 또는 타입
        투표구명 = 타입_val or emd_val
        emd_actual = emd_val if 타입_val else ""

        if is_skip_row(emd_val) and is_skip_row(타입_val):
            continue
        # 합계/소계 행 제거 (읍면동이 '합계', 투표타입이 없는 경우)
        if emd_val in ("합계",):
            continue
        if is_skip_row(투표구명):
            continue

        선거인수 = to_int(row[선거인수_col])
        투표수 = to_int(row[투표수_col])
        무효 = to_int(row[무효_col]) if 무효_col is not None and 무효_col < len(row) else None
        기권 = to_int(row[기권_col]) if 기권_col is not None and 기권_col < len(row) else None

        for c, party, cand in current_candidates:
            득표 = to_int(row[c]) if c < len(row) else None
            rows.append({
                "선거_회차": "제22대",
                "선거일": "2024-04-10",
                "선거구분": "지역구",
                "시도": current_sido,
                "구시군": current_구시군,
                "읍면동": emd_actual,
                "투표구": 투표구명,
                "선거구명": current_선거구,
                "선거인수": 선거인수,
                "투표수": 투표수,
                "후보자": cand,
                "정당": party,
                "득표수": 득표,
                "무효투표수": 무효,
                "기권수": 기권,
                "level": get_level(투표구명),
            })

    return pd.DataFrame(rows, columns=FINAL_COLS)


def parse_22대_비례():
    path = os.path.join(BASE, "data_raw", "개표결과", "2. 개표단위별 개표결과(비례대표) -전국.xlsx")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[SKIP] 22대 비례: {e}")
        return pd.DataFrame(columns=FINAL_COLS)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    rows = []

    if len(all_rows) < 7:
        return pd.DataFrame(columns=FINAL_COLS)

    # row4 (idx=3): '시도명', '구시군명', '읍면동명', '투표구명', '선거인수', '투표수', ...
    # row5 (idx=4): 정당명들
    header_row = list(all_rows[3])
    party_row = list(all_rows[4])
    ncols = len(header_row)

    시도_col = 0
    구시군_col = 1
    emd_col = 2
    투표구_col = 3
    선거인수_col = 4
    투표수_col = 5
    cand_start = 6

    무효_col = None
    기권_col = None
    for c in range(ncols - 1, -1, -1):
        v = str(header_row[c] or "").replace("\n", "").strip()
        if "기권" in v and 기권_col is None:
            기권_col = c
        if "무효" in v and 무효_col is not None:
            pass
        if "무효" in v and 무효_col is None:
            무효_col = c

    end_col = min(c for c in [무효_col, 기권_col] if c is not None) if (무효_col or 기권_col) else ncols

    candidate_cols = []
    for c in range(cand_start, end_col):
        party = str(party_row[c] or "").strip()
        if party in ("계", "", "None"):
            continue
        candidate_cols.append((c, party))

    current_sido = None
    current_구시군 = None

    for r_idx in range(6, len(all_rows)):
        row = all_rows[r_idx]
        if len(row) < 6:
            continue

        sido_val = str(row[시도_col] or "").strip()
        구시군_val = str(row[구시군_col] or "").strip()
        emd_val = str(row[emd_col] or "").strip()
        투표구명 = str(row[투표구_col] or "").strip()

        if sido_val:
            current_sido = normalize_sido(sido_val)
        if 구시군_val:
            current_구시군 = 구시군_val

        # 합계/소계 제거
        if is_skip_row(emd_val) and is_skip_row(투표구명):
            continue
        if is_skip_row(투표구명) and not emd_val:
            continue
        if is_skip_row(투표구명):
            continue

        if not current_sido:
            continue

        선거인수 = to_int(row[선거인수_col])
        투표수 = to_int(row[투표수_col])
        무효 = to_int(row[무효_col]) if 무효_col is not None and 무효_col < len(row) else None
        기권 = to_int(row[기권_col]) if 기권_col is not None and 기권_col < len(row) else None

        for c, party in candidate_cols:
            득표 = to_int(row[c]) if c < len(row) else None
            rows.append({
                "선거_회차": "제22대",
                "선거일": "2024-04-10",
                "선거구분": "비례대표",
                "시도": current_sido,
                "구시군": current_구시군,
                "읍면동": emd_val,
                "투표구": 투표구명,
                "선거구명": "",
                "선거인수": 선거인수,
                "투표수": 투표수,
                "후보자": "",
                "정당": party,
                "득표수": 득표,
                "무효투표수": 무효,
                "기권수": 기권,
                "level": get_level(투표구명),
            })

    return pd.DataFrame(rows, columns=FINAL_COLS)


# ─────────────────────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("파싱 시작...")
    all_dfs = []

    print("  16대 파싱...")
    df16 = parse_16대()
    print(f"    → {len(df16):,}행")
    all_dfs.append(df16)

    print("  17대 파싱...")
    df17 = parse_17대()
    print(f"    → {len(df17):,}행")
    all_dfs.append(df17)

    print("  18대 파싱...")
    df18 = parse_18대()
    print(f"    → {len(df18):,}행")
    all_dfs.append(df18)

    print("  19대 파싱...")
    df19 = parse_19대()
    print(f"    → {len(df19):,}행")
    all_dfs.append(df19)

    print("  20대 파싱...")
    df20 = parse_20대()
    print(f"    → {len(df20):,}행")
    all_dfs.append(df20)

    print("  21대 파싱...")
    df21 = parse_21대()
    print(f"    → {len(df21):,}행")
    all_dfs.append(df21)

    print("  22대 지역구 파싱...")
    df22j = parse_22대_지역구()
    print(f"    → {len(df22j):,}행")
    all_dfs.append(df22j)

    print("  22대 비례 파싱...")
    df22b = parse_22대_비례()
    print(f"    → {len(df22b):,}행")
    all_dfs.append(df22b)

    print("  통합 중...")
    df = pd.concat(all_dfs, ignore_index=True)

    # 숫자 컬럼 int 변환 (nullable int)
    for col in ["선거인수", "투표수", "득표수", "무효투표수", "기권수"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

    # 불필요한 빈값 행 제거: 투표구, 후보자/정당 둘 다 없는 행
    df = df[~(df["투표구"].eq("") & df["정당"].eq("") & df["후보자"].eq(""))]

    os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
    df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

    print(f"\n저장 완료: {OUT_CSV}")
    print(f"총 행수: {len(df):,}")
    print(f"컬럼: {list(df.columns)}")
    print("\n샘플 5행:")
    print(df.head(5).to_string())


if __name__ == "__main__":
    main()
