"""
전국동시지방선거 raw 데이터 → 통합 CSV 변환 스크립트
출력: data_processed/지방선거.csv
"""

import hashlib
import os
import re
import unicodedata
import xlrd
import openpyxl
import pandas as pd
import warnings

warnings.filterwarnings("ignore")

BASE = "/Users/seong-in/Desktop/Git/pyosim/data_raw"
OUT_DIR = "/Users/seong-in/Desktop/Git/pyosim/data_processed"
os.makedirs(OUT_DIR, exist_ok=True)

ELECTION_DATES = {3: "2002-06-13", 4: "2006-05-31", 5: "2010-06-02",
                  6: "2014-06-04", 7: "2018-06-13", 8: "2022-06-01"}

FINAL_COLS = ["선거_회차", "선거일", "선거종류", "시도", "구시군", "읍면동",
              "선거구명", "선거인수", "투표수", "후보자", "정당", "득표수",
              "무효투표수", "기권수", "level"]

ALL_ROWS = []


# ─────────────────────────────────────────────────────────────────────────────
# 공통 유틸
# ─────────────────────────────────────────────────────────────────────────────
def clean_num(value):
    """쉼표 제거 + 숫자 변환. 변환 불가면 None."""
    if value is None or (isinstance(value, str) and value.strip() in ("", "-")):
        return None
    try:
        return int(str(value).replace(",", "").strip().split(".")[0])
    except (ValueError, AttributeError):
        return None


def split_party_name(header: str):
    """'정당명\\n후보자명' 또는 '정당명 후보자명' → (정당, 후보자)."""
    if header is None:
        return None, None
    header = str(header).replace("_x000D_", "").strip()
    if "\n" in header:
        parts = header.split("\n", 1)
        return parts[0].strip(), parts[1].strip()
    # 공백으로 분리 (첫 토큰 = 정당)
    parts = header.strip().split(" ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return header.strip(), ""


LEVEL_MAP = {
    "관내사전투표": "사전투표",
    "관외사전투표": "관외사전투표",
    "거소": "거소선상",
    "거소우편투표": "거소선상",
    "거소투표": "거소선상",
    "거소·선상투표": "거소선상",
    "선거일투표": "당일투표",
}


def get_level(name: str) -> str:
    s = str(name).strip() if name else ""
    return LEVEL_MAP.get(s, "당일투표")


def is_summary_row(name: str) -> bool:
    """합계, 소계, 부재자, 잘못투입 등 제거 대상."""
    if not name:
        return False
    skip_keywords = ["합계", "소계", "부재자", "잘못투입된투표지", "잘못투입", "계"]
    for kw in skip_keywords:
        if str(name).strip() == kw:
            return True
    return False


def emit_row(round_num, election_type, sido, gu, emd, district, voters,
             votes, candidate, party, score, invalid, abstain, level="당일투표"):
    ALL_ROWS.append({
        "선거_회차": round_num,
        "선거일": ELECTION_DATES[round_num],
        "선거종류": election_type,
        "시도": sido or "",
        "구시군": gu or "",
        "읍면동": emd or "",
        "선거구명": district or "",
        "선거인수": clean_num(voters),
        "투표수": clean_num(votes),
        "후보자": candidate or "",
        "정당": party or "",
        "득표수": clean_num(score),
        "무효투표수": clean_num(invalid),
        "기권수": clean_num(abstain),
        "level": level,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 3회 파일 파서 (xls)
# 구조: 위원회명 | 투표구명 | 선거인수 | 투표수 | 후보1 | 후보2 ... | 계 | ...
# Header row 2 (0-indexed) contains candidate names starting at col 4
# ─────────────────────────────────────────────────────────────────────────────
GU_SIDO_3RD = {
    # 서울
    "종로구": "서울", "중구": "서울", "용산구": "서울", "성동구": "서울", "광진구": "서울",
    "동대문구": "서울", "중랑구": "서울", "성북구": "서울", "강북구": "서울", "도봉구": "서울",
    "노원구": "서울", "은평구": "서울", "서대문구": "서울", "마포구": "서울", "양천구": "서울",
    "강서구": "서울", "구로구": "서울", "금천구": "서울", "영등포구": "서울", "동작구": "서울",
    "관악구": "서울", "서초구": "서울", "강남구": "서울", "송파구": "서울", "강동구": "서울",
    # 부산
    "중구(부산)": "부산", "서구(부산)": "부산", "동구(부산)": "부산", "영도구": "부산",
    "부산진구": "부산", "동래구": "부산", "남구(부산)": "부산", "북구(부산)": "부산",
    "해운대구": "부산", "사하구": "부산", "금정구": "부산", "강서구(부산)": "부산",
    "연제구": "부산", "수영구": "부산", "사상구": "부산", "기장군": "부산",
    # 대구
    "중구(대구)": "대구", "동구(대구)": "대구", "서구(대구)": "대구", "남구(대구)": "대구",
    "북구(대구)": "대구", "수성구": "대구", "달서구": "대구", "달성군": "대구",
    # 인천
    "중구(인천)": "인천", "동구(인천)": "인천", "남구(인천)": "인천", "연수구": "인천",
    "남동구": "인천", "부평구": "인천", "계양구": "인천", "서구(인천)": "인천",
    "강화군": "인천", "옹진군": "인천",
    # 광주
    "동구(광주)": "광주", "서구(광주)": "광주", "남구(광주)": "광주",
    "북구(광주)": "광주", "광산구": "광주",
    # 대전
    "동구(대전)": "대전", "중구(대전)": "대전", "서구(대전)": "대전",
    "유성구": "대전", "대덕구": "대전",
    # 울산
    "중구(울산)": "울산", "남구(울산)": "울산", "동구(울산)": "울산",
    "북구(울산)": "울산", "울주군": "울산",
    # 경기
    "수원시": "경기", "성남시": "경기", "의정부시": "경기", "안양시": "경기",
    "부천시": "경기", "광명시": "경기", "평택시": "경기", "동두천시": "경기",
    "안산시": "경기", "고양시": "경기", "과천시": "경기", "구리시": "경기",
    "남양주시": "경기", "오산시": "경기", "시흥시": "경기", "군포시": "경기",
    "의왕시": "경기", "하남시": "경기", "용인시": "경기", "파주시": "경기",
    "이천시": "경기", "안성시": "경기", "김포시": "경기", "화성시": "경기",
    "광주시": "경기", "양주시": "경기", "여주군": "경기", "연천군": "경기",
    "포천군": "경기", "가평군": "경기", "양평군": "경기",
    # 강원
    "춘천시": "강원", "원주시": "강원", "강릉시": "강원", "동해시": "강원",
    "태백시": "강원", "속초시": "강원", "삼척시": "강원", "홍천군": "강원",
    "횡성군": "강원", "영월군": "강원", "평창군": "강원", "정선군": "강원",
    "철원군": "강원", "화천군": "강원", "양구군": "강원", "인제군": "강원",
    "고성군": "강원", "양양군": "강원",
    # 충북
    "청주시": "충북", "충주시": "충북", "제천시": "충북", "청원군": "충북",
    "보은군": "충북", "옥천군": "충북", "영동군": "충북", "진천군": "충북",
    "괴산군": "충북", "음성군": "충북", "단양군": "충북", "증평군": "충북",
    # 충남
    "천안시": "충남", "공주시": "충남", "보령시": "충남", "아산시": "충남",
    "서산시": "충남", "논산시": "충남", "계룡시": "충남", "금산군": "충남",
    "연기군": "충남", "부여군": "충남", "서천군": "충남", "청양군": "충남",
    "홍성군": "충남", "예산군": "충남", "태안군": "충남", "당진군": "충남",
    # 전북
    "전주시": "전북", "군산시": "전북", "익산시": "전북", "정읍시": "전북",
    "남원시": "전북", "김제시": "전북", "완주군": "전북", "진안군": "전북",
    "무주군": "전북", "장수군": "전북", "임실군": "전북", "순창군": "전북",
    "고창군": "전북", "부안군": "전북",
    # 전남
    "목포시": "전남", "여수시": "전남", "순천시": "전남", "나주시": "전남",
    "광양시": "전남", "담양군": "전남", "곡성군": "전남", "구례군": "전남",
    "고흥군": "전남", "보성군": "전남", "화순군": "전남", "장흥군": "전남",
    "강진군": "전남", "해남군": "전남", "영암군": "전남", "무안군": "전남",
    "함평군": "전남", "영광군": "전남", "장성군": "전남", "완도군": "전남",
    "진도군": "전남", "신안군": "전남",
    # 경북
    "포항시": "경북", "경주시": "경북", "김천시": "경북", "안동시": "경북",
    "구미시": "경북", "영주시": "경북", "영천시": "경북", "상주시": "경북",
    "문경시": "경북", "경산시": "경북", "군위군": "경북", "의성군": "경북",
    "청송군": "경북", "영양군": "경북", "영덕군": "경북", "청도군": "경북",
    "고령군": "경북", "성주군": "경북", "칠곡군": "경북", "예천군": "경북",
    "봉화군": "경북", "울진군": "경북", "울릉군": "경북",
    # 경남
    "창원시": "경남", "마산시": "경남", "진주시": "경남", "진해시": "경남",
    "통영시": "경남", "사천시": "경남", "김해시": "경남", "밀양시": "경남",
    "거제시": "경남", "양산시": "경남", "의령군": "경남", "함안군": "경남",
    "창녕군": "경남", "고성군(경남)": "경남", "남해군": "경남", "하동군": "경남",
    "산청군": "경남", "함양군": "경남", "거창군": "경남", "합천군": "경남",
    # 제주
    "제주시": "제주", "서귀포시": "제주", "북제주군": "제주", "남제주군": "제주",
}


def _guess_gu_from_xls(ws) -> str:
    """3회 xls 파일 내용에서 구시군명 추론 (시도의원 파일의 선거구명 셀 활용)."""
    for row_idx in range(min(8, ws.nrows)):
        for col_idx in range(min(4, ws.ncols)):
            val = str(ws.cell_value(row_idx, col_idx)).strip()
            # "오산시제1선거구" 같은 패턴에서 구시군명 추출
            m = re.match(r"^([가-힣]+시|[가-힣]+군|[가-힣]+구)", val)
            if m:
                candidate = m.group(1)
                if candidate in GU_SIDO_3RD:
                    return candidate
    return ""


def parse_3rd_xls(filepath, election_type, sido_from_filename):
    try:
        wb = xlrd.open_workbook(filepath, encoding_override="cp949")
    except Exception:
        wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    # 파일명이 한글이 아닌 경우(깨진 파일명) 내용에서 구시군 추론
    is_valid_gu = bool(re.match(r"^[가-힣]", sido_from_filename))
    if not is_valid_gu:
        guessed = _guess_gu_from_xls(ws)
        sido_from_filename = guessed if guessed else sido_from_filename

    # Find candidate header row (contains party/candidate names)
    cand_row_idx = None
    for row_idx in range(min(5, ws.nrows)):
        row_vals = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
        non_empty = [str(v).strip() for v in row_vals if str(v).strip()]
        # Candidate row has multiple non-empty cells excluding standard headers
        std_headers = {"위원회명", "투표구명", "선거 인수", "투표수", "계",
                       "투 표 율 (%)", "유효 투표 율 (%)", "무효 투표 수", "기 권 수",
                       "무효투표수", "기권수"}
        candidates_found = [v for v in non_empty if v not in std_headers
                            and "유효투표수" not in v and "후보자별" not in v
                            and len(v) > 1]
        if len(candidates_found) >= 2:
            cand_row_idx = row_idx
            break

    if cand_row_idx is None:
        return

    cand_headers = [ws.cell_value(cand_row_idx, col) for col in range(ws.ncols)]

    # Find where candidates start/end
    # Candidate cols start after 투표수 (col 3) and end before '계'
    # Col 0: 위원회명(시군구), Col 1: 투표구명, Col 2: 선거인수, Col 3: 투표수
    # Col 4+: candidates until blank or '계'

    # Find fixed columns
    invalid_col = None
    abstain_col = None
    total_col = None
    for col_idx, hdr in enumerate(cand_headers):
        hdr_str = str(hdr).strip()
        if hdr_str in ("무효 투표 수", "무효투표수", "무효\n투표수"):
            invalid_col = col_idx
        elif hdr_str in ("기 권 수", "기권수", "기권\n수"):
            abstain_col = col_idx
        elif hdr_str == "계":
            total_col = col_idx

    # 3회 파일은 두 가지 구조:
    #   A형(시도지사/광역비례): col0=위원회명, col1=투표구명, col2=선거인수, col3=투표수, col4=유효, col5+=후보
    #   B형(구시군장/시도의원/구시군의원): col0=투표구명, col1=선거인수, col2=투표수, col3=유효, col4+=후보
    row0_col0 = str(ws.cell_value(0, 0)).strip()
    has_wiwon_col = (row0_col0 == "위원회명")

    end_col = total_col or invalid_col or ws.ncols
    cand_cols = []
    cand_start = 5 if has_wiwon_col else 4
    for col_idx in range(cand_start, end_col):
        hdr = str(cand_headers[col_idx]).strip()
        if hdr and hdr not in ("", "계", "유효투표수", "유효 투표수", "후보자별 득표수"):
            cand_cols.append((col_idx, hdr))

    # Parse data rows
    current_sido = sido_from_filename
    current_gu = sido_from_filename  # B형 기본값 (파일명=구시군명)
    for row_idx in range(cand_row_idx + 1, ws.nrows):
        col0 = str(ws.cell_value(row_idx, 0)).strip()

        if has_wiwon_col:
            # A형: col0=위원회명(구시군), col1=투표구명
            if col0 and col0 not in ("", "0.0"):
                current_gu = re.sub(r"\(.+\)", "", col0).strip()
                if current_gu in GU_SIDO_3RD:
                    current_sido = GU_SIDO_3RD[current_gu]
            emd = str(ws.cell_value(row_idx, 1)).strip()
            if is_summary_row(emd) or is_summary_row(col0) or not emd or emd in ("", "0.0"):
                continue
            voters = ws.cell_value(row_idx, 2)
            votes = ws.cell_value(row_idx, 3)
        else:
            # B형: col0=투표구명
            if is_summary_row(col0) or not col0 or col0 in ("", "0.0"):
                continue
            emd = col0
            voters = ws.cell_value(row_idx, 1)
            votes = ws.cell_value(row_idx, 2)

        invalid = ws.cell_value(row_idx, invalid_col) if invalid_col is not None else None
        abstain = ws.cell_value(row_idx, abstain_col) if abstain_col is not None else None

        for col_idx, cand_header in cand_cols:
            score = ws.cell_value(row_idx, col_idx)
            party, candidate = split_party_name(cand_header)
            if clean_num(score) is None:
                continue
            emit_row(3, election_type, current_sido, current_gu, emd,
                     current_gu, voters, votes, candidate, party, score, invalid, abstain,
                     level=get_level(emd))


# ─────────────────────────────────────────────────────────────────────────────
# 4회 파일 파서 (xls)
# 구조: Row 3: 읍면동명, 선거인수, 투표수, 후보자별 득표수...
#       Row 4: 후보자명 (정당\n후보자)
#       Row 1: [시도명] 있음
# ─────────────────────────────────────────────────────────────────────────────
def parse_4th_xls(filepath, election_type, sido_hint, gu_hint):
    try:
        wb = xlrd.open_workbook(filepath, encoding_override="cp949")
    except Exception:
        wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    # Find header structure - scan for row with 읍면동명 or 선거인수
    header_row = None
    cand_row = None
    for row_idx in range(min(8, ws.nrows)):
        row_vals = [str(ws.cell_value(row_idx, col)).strip() for col in range(ws.ncols)]
        if "읍면동명" in row_vals or "선거인수" in row_vals:
            header_row = row_idx
            # Candidate row is next non-empty row
            for next_row in range(row_idx + 1, min(row_idx + 3, ws.nrows)):
                next_vals = [str(ws.cell_value(next_row, col)).strip() for col in range(ws.ncols)]
                non_empty = [v for v in next_vals if v]
                if non_empty:
                    cand_row = next_row
                    break
            break

    if header_row is None:
        return

    header_vals = [str(ws.cell_value(header_row, col)).strip() for col in range(ws.ncols)]
    cand_vals = [str(ws.cell_value(cand_row, col)).strip() for col in range(ws.ncols)] if cand_row else []

    # Try to extract sido from file - look in rows 0-2
    sido = sido_hint
    gu = gu_hint
    for row_idx in range(min(3, ws.nrows)):
        row_str = " ".join(str(ws.cell_value(row_idx, col)) for col in range(ws.ncols))
        # Look for [시도명] pattern
        match = re.search(r"\[([가-힣]+(?:특별|광역|자치)?(?:시|도|특별자치시|특별자치도))\]", row_str)
        if match:
            sido = match.group(1)
        # Look for [구시군명] pattern
        match2 = re.search(r"\[([가-힣\s]+(?:구|시|군|특별자치도))\]", row_str)
        if match2 and match2.group(1) != sido:
            gu = match2.group(1).strip()

    # Find columns
    emd_col = None
    voters_col = None
    votes_col = None
    invalid_col = None
    abstain_col = None
    total_col = None

    for col_idx, hdr in enumerate(header_vals):
        if hdr in ("읍면동명", "읍 면 동 명"):
            emd_col = col_idx
        elif hdr in ("선거인수", "선 거 인 수"):
            voters_col = col_idx
        elif hdr in ("투표수", "투 표 수"):
            votes_col = col_idx
        elif "무효" in hdr:
            invalid_col = col_idx
        elif "기권" in hdr:
            abstain_col = col_idx
        elif hdr == "계":
            total_col = col_idx

    if emd_col is None:
        # Some files: first col is 읍면동명 without explicit header
        emd_col = 1
        voters_col = 2
        votes_col = 4

    # Find candidate columns from cand_row
    # Candidates are between votes_col+1 and total_col
    start_col = (votes_col or 4) + 1
    end_col = total_col or invalid_col or ws.ncols
    cand_cols = []
    for col_idx in range(start_col, end_col):
        cand_hdr = cand_vals[col_idx] if cand_row and col_idx < len(cand_vals) else ""
        if cand_hdr.strip() and cand_hdr.strip() != "계":
            cand_cols.append((col_idx, cand_hdr))

    # Also check header_vals for candidates if cand_vals didn't work
    if not cand_cols and header_vals:
        for col_idx in range(start_col, end_col):
            hdr = header_vals[col_idx]
            if hdr and hdr not in ("계", "후보자별 득표수", "유효 투표수"):
                cand_cols.append((col_idx, hdr))

    current_gu = gu
    current_district = gu

    for row_idx in range((cand_row or header_row) + 1, ws.nrows):
        vals = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
        emd = str(vals[emd_col]).strip() if emd_col is not None and emd_col < len(vals) else ""

        if not emd or emd in ("0.0", ""):
            continue
        if is_summary_row(emd):
            continue

        # Check if this row updates gu (시군구 변경 표시)
        # In some 4th election files, column 0 might be blank or hold gu
        col0 = str(vals[0]).strip() if vals else ""
        if col0 and col0 not in ("", "0.0") and col0 != emd:
            potential_gu = re.sub(r"\(.+\)", "", col0).strip()
            if potential_gu:
                current_gu = potential_gu

        voters = vals[voters_col] if voters_col is not None else None
        votes = vals[votes_col] if votes_col is not None else None
        invalid = vals[invalid_col] if invalid_col is not None else None
        abstain = vals[abstain_col] if abstain_col is not None else None

        for col_idx, cand_header in cand_cols:
            if col_idx >= len(vals):
                continue
            score = vals[col_idx]
            party, candidate = split_party_name(cand_header)
            if clean_num(score) is None:
                continue
            emit_row(4, election_type, sido, current_gu, emd,
                     current_gu, voters, votes, candidate, party, score, invalid, abstain,
                     level=get_level(emd))


# ─────────────────────────────────────────────────────────────────────────────
# 5회/6회 파일 파서 (xls/xlsx)
# 구조:
#   Row 0: 개표진행상황(읍면동별)
#   Row 1: blank
#   Row 2: [선거종류][시도명]  or  [선거종류][시도명][구시군명]
#   Row 3: 구시군명 | 읍면동명 | 선거인수 | 투표수 | 후보자별 득표수 | ... | 무효투표수 | 기권수
#   Row 4: | | | | 후보1 | 후보2 | ... | 계 | |
#   Row 5+: 데이터
# ─────────────────────────────────────────────────────────────────────────────
def parse_5th_6th_file(filepath, round_num, election_type, sido_hint=""):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        try:
            wb = xlrd.open_workbook(filepath, encoding_override="cp949")
        except Exception:
            wb = xlrd.open_workbook(filepath)
        sheets = [wb.sheet_by_index(i) for i in range(wb.nsheets)]
    else:
        wb_xlsx = openpyxl.load_workbook(filepath, data_only=True)
        sheets = list(wb_xlsx.worksheets)

    for ws in sheets:
        if ext == ".xls":
            get_cell = lambda r, c: ws.cell_value(r, c)
            nrows = ws.nrows
            ncols = ws.ncols
        else:
            all_rows = list(ws.iter_rows(values_only=True))
            nrows = len(all_rows)
            ncols = ws.max_column
            get_cell = lambda r, c: all_rows[r][c] if c < len(all_rows[r]) else None

        # Extract sido and election type from header rows
        sido = sido_hint
        gu_from_header = ""
        for scan_row in range(min(4, nrows)):
            row_str = " ".join(str(get_cell(scan_row, col) or "") for col in range(min(ncols, 20)))
            # Match [시도명] or [시·도지사선거][시도명]
            matches = re.findall(r"\[([^\]]+)\]", row_str)
            for match in matches:
                match = match.strip()
                if re.search(r"(선거|투표|소)", match):
                    continue
                if re.search(r"(특별자치시|특별자치도|광역시|특별시|도)$", match) or \
                   re.search(r"^(서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)", match):
                    if "구" not in match and "군" not in match and "시장" not in match:
                        if len(match) < 15:
                            sido = match
                if re.search(r"(구|시|군)$", match) and not re.search(r"(선거|의원|의장|감)", match):
                    gu_from_header = match

        # Find the column header row (구시군명 or 읍면동명)
        col_header_row = None
        for row_idx in range(min(6, nrows)):
            row_vals = [str(get_cell(row_idx, col) or "").strip() for col in range(ncols)]
            if any(v in ("구시군명", "읍면동명", "시도명", "위원회명") for v in row_vals):
                col_header_row = row_idx
                break

        if col_header_row is None:
            continue

        header_vals = [str(get_cell(col_header_row, col) or "").strip() for col in range(ncols)]

        # Candidate header row = next row that contains actual candidate names
        # Skip rows that contain "후보자별" or other header-only text
        cand_row_idx = None
        for candidate_scan in range(col_header_row + 1, min(col_header_row + 4, nrows)):
            scan_vals = [str(get_cell(candidate_scan, col) or "").strip() for col in range(ncols)]
            scan_non_empty = [v for v in scan_vals if v and v not in ("계", "후보자별 득표수")]
            # A row with "후 보 자 별 득 표 수" or "정 당 별 득 표 수" (spaced) is NOT a candidate row
            combined = " ".join(scan_non_empty)
            if "후 보 자 별" in combined or "보 자 별 득" in combined:
                continue
            if "당 별 득 표" in combined or "정 당 별" in combined:
                continue
            # Check if any cell looks like a candidate (contains \n or has 당 in it)
            has_candidates = any(
                ("\n" in v or "당" in v or "무소속" in v or len(v) >= 2)
                for v in scan_non_empty
                if v not in ("구분", "시도명", "구시군명", "읍면동명", "선거인수", "투표수", "계")
            )
            if has_candidates:
                cand_row_idx = candidate_scan
                break
        if cand_row_idx is None:
            cand_row_idx = col_header_row + 1
        if cand_row_idx >= nrows:
            continue
        cand_vals = [str(get_cell(cand_row_idx, col) or "").strip() for col in range(ncols)]

        # Identify columns
        sido_col = None
        gu_col = None
        emd_col = None
        voters_col = None
        votes_col = None
        invalid_col = None
        abstain_col = None
        total_col = None
        district_col = None

        for col_idx, hdr in enumerate(header_vals):
            if hdr in ("시도명", "시·도명"):
                sido_col = col_idx
            elif hdr in ("구시군명", "위원회명"):
                gu_col = col_idx
            elif hdr in ("읍면동명", "투표구명"):
                emd_col = col_idx
            elif "선거인수" in hdr:
                voters_col = col_idx
            elif hdr in ("투표수",) and "유효" not in hdr:
                votes_col = col_idx
            elif "무효" in hdr:
                invalid_col = col_idx
            elif "기권" in hdr:
                abstain_col = col_idx
            elif hdr in ("계",):
                total_col = col_idx
            elif "선거구" in hdr:
                district_col = col_idx

        # Fallbacks
        if emd_col is None:
            emd_col = 1 if gu_col == 0 else 0
        if voters_col is None:
            voters_col = (emd_col or 0) + 1
        if votes_col is None:
            votes_col = voters_col + 1

        # Find candidate columns
        start_col = votes_col + 1
        end_col = total_col or invalid_col or ncols
        cand_cols = []
        for col_idx in range(start_col, end_col):
            # Check both header row and cand row
            hdr_cand = cand_vals[col_idx] if col_idx < len(cand_vals) else ""
            if hdr_cand and hdr_cand not in ("계", "후보자별 득표수", "", "유 효 투 표 수 "):
                cand_cols.append((col_idx, hdr_cand))

        # If still no candidates found, check header_vals
        if not cand_cols:
            for col_idx in range(start_col, end_col):
                hdr = header_vals[col_idx] if col_idx < len(header_vals) else ""
                if hdr and hdr not in ("계", "후보자별 득표수", ""):
                    cand_cols.append((col_idx, hdr))

        current_gu = gu_from_header
        current_district = gu_from_header
        data_start = cand_row_idx + 1

        # Skip blank rows after cand row
        while data_start < nrows:
            test_vals = [get_cell(data_start, col) for col in range(min(5, ncols))]
            if any(v is not None and str(v).strip() not in ("", "0", "0.0") for v in test_vals):
                break
            data_start += 1

        for row_idx in range(data_start, nrows):
            vals = [get_cell(row_idx, col) for col in range(ncols)]

            col0_val = str(vals[0] or "").strip() if vals else ""
            gu_val = str(vals[gu_col] or "").strip() if gu_col is not None and gu_col < len(vals) else ""
            emd_val = str(vals[emd_col] or "").strip() if emd_col is not None and emd_col < len(vals) else ""
            district_val = str(vals[district_col] or "").strip() if district_col is not None and district_col < len(vals) else ""

            # Update current_gu
            if gu_col is not None and gu_val and not is_summary_row(gu_val):
                current_gu = gu_val
            elif gu_col is None and col0_val and not is_summary_row(col0_val):
                current_gu = col0_val

            if district_val:
                current_district = district_val

            # Skip summary/empty
            if not emd_val or emd_val in ("0.0", ""):
                continue
            if is_summary_row(emd_val):
                continue

            # 구분 컬럼으로 level 결정 (소계만 스킵, 관내사전/선거일은 살림)
            구분_col = emd_col + 1 if emd_col is not None else None
            row_level = "당일투표"
            if 구분_col is not None and 구분_col < len(vals):
                구분 = str(vals[구분_col] or "").strip()
                if 구분 in ("관내", "관외"):
                    continue
                row_level = get_level(구분) if 구분 else get_level(emd_val)

            # Use sido_col if available
            row_sido = sido
            if sido_col is not None and sido_col < len(vals):
                sv = str(vals[sido_col] or "").strip()
                if sv:
                    row_sido = sv

            voters = vals[voters_col] if voters_col is not None and voters_col < len(vals) else None
            votes = vals[votes_col] if votes_col is not None and votes_col < len(vals) else None
            invalid = vals[invalid_col] if invalid_col is not None and invalid_col < len(vals) else None
            abstain = vals[abstain_col] if abstain_col is not None and abstain_col < len(vals) else None

            for col_idx, cand_header in cand_cols:
                if col_idx >= len(vals):
                    continue
                score = vals[col_idx]
                party, candidate = split_party_name(cand_header)
                if clean_num(score) is None:
                    continue
                emit_row(round_num, election_type, row_sido, current_gu, emd_val,
                         current_district or current_gu, voters, votes,
                         candidate, party, score, invalid, abstain,
                         level=row_level)


# ─────────────────────────────────────────────────────────────────────────────
# 7회 파일 파서 (xlsx)
# 구조: Row 0: 선거종류 | 선거구명 | 시도명 | 구시군명 | 읍면동명 | 구분 | 선거인수 | 투표수 | 후보자별... | 무효투표수 | 기권수
#       Row 1: (candidate names merged from row 0 span)
#       Row 2+: data
# ─────────────────────────────────────────────────────────────────────────────
def parse_7th_xlsx(filepath, election_type):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    row0 = list(all_rows[0])
    row1 = list(all_rows[1])

    # Col positions from row0
    sido_col = None
    gu_col = None
    emd_col = None
    district_col = None
    voters_col = None
    votes_col = None
    invalid_col = None
    abstain_col = None

    for col_idx, hdr in enumerate(row0):
        hdr_str = str(hdr or "").strip()
        if hdr_str == "시도명":
            sido_col = col_idx
        elif hdr_str == "구시군명":
            gu_col = col_idx
        elif hdr_str == "읍면동명":
            emd_col = col_idx
        elif hdr_str == "선거구명":
            district_col = col_idx
        elif hdr_str == "선거인수":
            voters_col = col_idx
        elif hdr_str == "투표수":
            votes_col = col_idx
        elif hdr_str == "무효투표수":
            invalid_col = col_idx
        elif hdr_str == "기권수":
            abstain_col = col_idx

    if voters_col is None:
        return

    # Candidate cols: between votes_col and invalid_col (from row1)
    start_col = votes_col + 1
    end_col = invalid_col or len(row1)
    cand_cols = []
    for col_idx in range(start_col, end_col):
        cand_name = str(row1[col_idx] or "").strip()
        if cand_name and cand_name not in ("계", ""):
            cand_cols.append((col_idx, cand_name))

    # Data starts from row 2
    for row_vals in all_rows[2:]:
        row_vals = list(row_vals)
        # Pad
        while len(row_vals) < len(row0):
            row_vals.append(None)

        emd = str(row_vals[emd_col] or "").strip() if emd_col is not None else ""
        gu = str(row_vals[gu_col] or "").strip() if gu_col is not None else ""
        sido = str(row_vals[sido_col] or "").strip() if sido_col is not None else ""
        district = str(row_vals[district_col] or "").strip() if district_col is not None else ""

        # Skip summary rows
        if not emd or is_summary_row(emd):
            continue
        # 구분 컬럼으로 level 결정
        구분_col_7 = 5  # fixed position for 7th
        row_level_7 = "당일투표"
        if 구분_col_7 < len(row_vals):
            구분 = str(row_vals[구분_col_7] or "").strip()
            if 구분 in ("관내", "관외"):
                continue
            row_level_7 = get_level(구분) if 구분 else get_level(emd)

        voters = row_vals[voters_col] if voters_col is not None else None
        votes = row_vals[votes_col] if votes_col is not None else None
        invalid = row_vals[invalid_col] if invalid_col is not None else None
        abstain = row_vals[abstain_col] if abstain_col is not None else None

        for col_idx, cand_header in cand_cols:
            if col_idx >= len(row_vals):
                continue
            score = row_vals[col_idx]
            party, candidate = split_party_name(cand_header)
            if clean_num(score) is None:
                continue
            emit_row(7, election_type, sido, gu, emd, district or gu,
                     voters, votes, candidate, party, score, invalid, abstain,
                     level=row_level_7)


# ─────────────────────────────────────────────────────────────────────────────
# 8회 파일 파서 (xlsx)
# 구조: Row 0: 선거구명 or 시도명 | 구시군명 | 읍면동명 | 구분 | 선거인수 | 투표수 | 후보자별... | 계 | 무효투표수 | 기권수
#       Row 1: 정당1 | 정당2 | ... (for 비례)
#       Row 2: 후보자 이름 (정당\n후보자 or 정당 only for 비례)
#       Row 3+: data
# ─────────────────────────────────────────────────────────────────────────────
def parse_8th_xlsx(filepath, election_type):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) < 4:
        return

    row0 = list(all_rows[0])
    row1 = list(all_rows[1])
    row2 = list(all_rows[2])

    # Find column positions from row0
    sido_col = None
    gu_col = None
    emd_col = None
    district_col = None
    voters_col = None
    votes_col = None
    invalid_col = None
    abstain_col = None
    total_col = None

    for col_idx, hdr in enumerate(row0):
        hdr_str = str(hdr or "").strip()
        if hdr_str in ("시도명", "선거구명"):
            if sido_col is None:
                sido_col = col_idx
        elif hdr_str == "구시군명":
            gu_col = col_idx
        elif hdr_str == "읍면동명":
            emd_col = col_idx
        elif "선거인수" in hdr_str:
            voters_col = col_idx
        elif hdr_str == "투표수":
            votes_col = col_idx
        elif "무효" in hdr_str:
            invalid_col = col_idx
        elif "기권" in hdr_str:
            abstain_col = col_idx
        elif hdr_str == "계":
            total_col = col_idx

    if voters_col is None:
        return

    # Determine if it's a 비례 file (row2 has party names only, no \n candidate)
    # or regular file (row2 has party\n후보자 format)
    # Also some files use row2 for candidates (국회의원, 교육감)

    # The actual candidate headers are in row2
    start_col = votes_col + 1 if votes_col else 6
    end_col = total_col or invalid_col or len(row2)

    cand_cols = []
    for col_idx in range(start_col, end_col):
        # For 비례: row1 has party names, row2 repeats them
        # For candidates: row2 has party\n후보자
        cand_name_r2 = str(row2[col_idx] if col_idx < len(row2) else "").strip()
        cand_name_r1 = str(row1[col_idx] if col_idx < len(row1) else "").strip()

        # Prefer row2 (더 상세), fallback to row1
        cand_name = cand_name_r2 if cand_name_r2 and cand_name_r2 not in ("", "\n") else cand_name_r1
        if cand_name and cand_name not in ("계", ""):
            cand_cols.append((col_idx, cand_name))

    if not cand_cols:
        return

    current_sido = ""
    current_gu = ""
    current_district = ""

    # Data starts from row 3
    for row_vals in all_rows[3:]:
        row_vals = list(row_vals)
        while len(row_vals) < len(row0):
            row_vals.append(None)

        sido_val = str(row_vals[sido_col] or "").strip() if sido_col is not None else ""
        gu_val = str(row_vals[gu_col] or "").strip() if gu_col is not None else ""
        emd_val = str(row_vals[emd_col] or "").strip() if emd_col is not None else ""
        district_val = str(row_vals[district_col] or "").strip() if district_col is not None else ""

        # Update running values
        if sido_val:
            current_sido = sido_val
        if gu_val:
            current_gu = gu_val
        if district_val:
            current_district = district_val

        # Skip summaries
        if not emd_val or is_summary_row(emd_val):
            continue

        # 구분 컬럼으로 level 결정 (col 3 in 8th)
        구분_col_8 = 3
        row_level_8 = "당일투표"
        if 구분_col_8 < len(row_vals):
            구분 = str(row_vals[구분_col_8] or "").strip()
            if 구분 in ("관내", "관외"):
                continue
            row_level_8 = get_level(구분) if 구분 else get_level(emd_val)

        voters = row_vals[voters_col] if voters_col is not None else None
        votes = row_vals[votes_col] if votes_col is not None else None
        invalid = row_vals[invalid_col] if invalid_col is not None else None
        abstain = row_vals[abstain_col] if abstain_col is not None else None

        for col_idx, cand_header in cand_cols:
            if col_idx >= len(row_vals):
                continue
            score = row_vals[col_idx]
            party, candidate = split_party_name(cand_header)
            if clean_num(score) is None:
                continue
            emit_row(8, election_type, current_sido, current_gu, emd_val,
                     current_district or current_gu, voters, votes,
                     candidate, party, score, invalid, abstain,
                     level=row_level_8)


# ─────────────────────────────────────────────────────────────────────────────
# Directory traversal utilities
# ─────────────────────────────────────────────────────────────────────────────
def find_xls_files(dirpath):
    """Recursively find all xls/xlsx files."""
    result = []
    for root, dirs, files in os.walk(dirpath):
        for fname in files:
            if fname.lower().endswith((".xls", ".xlsx")) and not fname.startswith("~$"):
                result.append(os.path.join(root, fname))
    return result


# Mapping for election type standardization
def normalize_election_type(raw_type: str) -> str:
    mapping = {
        "시도지사": "시도지사", "도지사": "시도지사", "시장": "구시군장",
        "구시군장": "구시군장", "시장군수": "구시군장", "구시군의장": "구시군장",
        "시도의원": "시도의원", "도의원": "시도의원", "시의원": "시도의원",
        "지역구시도의원": "시도의원", "시도의회": "시도의원",
        "구시군의원": "구시군의원", "기초의원": "구시군의원", "지역구기초의원": "구시군의원",
        "구시군의회": "구시군의원",
        "광역비례": "광역비례", "비례대표시도의원": "광역비례", "광역의원비례대표": "광역비례",
        "비례대표광역": "광역비례", "광역비례의원": "광역비례",
        "기초비례": "기초비례", "비례대표기초의원": "기초비례", "기초의원비례대표": "기초비례",
        "비례대표기초": "기초비례", "기초비례의원": "기초비례",
        "교육감": "교육감", "교육의원": "교육의원",
        "국회의원": "국회의원재보궐", "국회의원재보궐": "국회의원재보궐",
    }
    for key, val in mapping.items():
        if key in raw_type:
            return val
    return raw_type


def infer_election_type_from_path(filepath: str) -> str:
    # Normalize to NFC to handle macOS NFD filenames
    filepath = unicodedata.normalize("NFC", filepath)
    fname = os.path.basename(filepath)
    parts = filepath.split(os.sep)

    # Check filename and parent directories
    for part in parts + [fname]:
        part_norm = part
        for kw, etype in [
            ("시도지사", "시도지사"), ("도지사", "시도지사"), ("구시군장", "구시군장"),
            ("구시군의장", "구시군장"), ("시장군수", "구시군장"),
            ("구군의장", "구시군장"), ("구청장", "구시군장"),
            ("시장", "구시군장"),
            ("시도의원", "시도의원"), ("도의원", "시도의원"), ("시도의회", "시도의원"),
            ("지역구시도", "시도의원"), ("시의원", "시도의원"),
            ("구시군의원", "구시군의원"), ("기초의원", "구시군의원"), ("구시군의회", "구시군의원"),
            ("지역구기초", "구시군의원"), ("시군의원", "구시군의원"), ("구의원", "구시군의원"),
            ("군의원", "구시군의원"),
            ("광역비례", "광역비례"), ("비례대표시도", "광역비례"), ("광역의원비례", "광역비례"),
            ("기초비례", "기초비례"), ("비례대표기초", "기초비례"), ("기초의원비례", "기초비례"),
            ("비례대표기초의원", "기초비례"),
            ("교육감", "교육감"), ("교육의원", "교육의원"),
            ("국회의원", "국회의원재보궐"),
        ]:
            if kw in part_norm:
                return etype
    return "기타"


# ─────────────────────────────────────────────────────────────────────────────
# 3회 processing
# ─────────────────────────────────────────────────────────────────────────────
def process_3rd():
    base = os.path.join(BASE, "전국동시지방선거 개표결과(제3회~제6회)", "제3회 전국동시지방선거 개표자료")
    type_dir_map = {
        "시도지사": "제3회지방선거시도지사",
        "광역비례": "제3회지방선거비례대표",
        "구시군장": "제3회지방선거구시군장",
        "시도의원": "제3회지방선거시도의원",
        "구시군의원": "제3회지방선거구시군의원",
    }
    for etype, dirname in type_dir_map.items():
        dirpath = os.path.join(base, dirname)
        if not os.path.isdir(dirpath):
            print(f"  [skip] {dirpath}")
            continue
        all_fnames = [f for f in os.listdir(dirpath) if f.lower().endswith(".xls")]
        # 파일 내용 해시로 중복 감지 (깨진 파일명 = 정상 파일의 복사본인 경우 스킵)
        # 한글 파일명을 비한글보다 먼저 처리해 정상 파일을 우선 등록
        seen_hashes: set = set()
        sorted_fnames = sorted(all_fnames, key=lambda f: (0 if re.match(r"^[가-힣]", f) else 1, f))
        for fname in sorted_fnames:
            fpath = os.path.join(dirpath, fname)
            file_hash = hashlib.md5(open(fpath, "rb").read()).hexdigest()
            if file_hash in seen_hashes:
                print(f"  [skip-dup] {fname} is duplicate of already-processed file")
                continue
            seen_hashes.add(file_hash)
            stem = os.path.splitext(fname)[0]
            try:
                parse_3rd_xls(fpath, etype, stem)
            except Exception as exc:
                print(f"  [error] 3rd {etype} {fname}: {exc}")
    print(f"  After 3rd: {len(ALL_ROWS)} rows")


# ─────────────────────────────────────────────────────────────────────────────
# 4회 processing
# ─────────────────────────────────────────────────────────────────────────────
SIDO_CODE_MAP = {
    "11": "서울특별시", "26": "부산광역시", "27": "대구광역시", "28": "인천광역시",
    "29": "광주광역시", "30": "대전광역시", "31": "울산광역시",
    "41": "경기도", "42": "강원도", "43": "충청북도", "44": "충청남도",
    "45": "전라북도", "46": "전라남도", "47": "경상북도", "48": "경상남도",
    "49": "제주특별자치도",
}


def sido_from_dirname(dirname: str) -> str:
    match = re.match(r"^(\d+)_", dirname)
    if match:
        code = match.group(1)
        return SIDO_CODE_MAP.get(code, dirname)
    return dirname


def process_4th():
    base = os.path.join(BASE, "전국동시지방선거 개표결과(제3회~제6회)", "제4회 전국동시지방선거 개표자료")
    type_dir_map = {
        "1_시도지사": "시도지사", "2_구시군장": "구시군장",
        "3_지역구시도의원": "시도의원", "4_지역구기초의원": "구시군의원",
        "6_비례대표시도의원": "광역비례", "7_비례대표기초의원": "기초비례",
        "8_교육의원": "교육의원",
    }
    for dirname, etype in type_dir_map.items():
        dirpath = os.path.join(base, dirname)
        if not os.path.isdir(dirpath):
            continue
        for fpath in sorted(find_xls_files(dirpath)):
            fname = os.path.basename(fpath)
            # Infer sido and gu from path
            rel_path = fpath.replace(dirpath, "").strip(os.sep)
            path_parts = rel_path.split(os.sep)
            sido = ""
            gu_hint = ""
            if len(path_parts) >= 2:
                sido = sido_from_dirname(path_parts[0])
            # gu from filename: CODE_구시군명_시도명.xls
            name_parts = os.path.splitext(fname)[0].split("_")
            if len(name_parts) >= 2:
                gu_hint = name_parts[1]
            try:
                parse_4th_xls(fpath, etype, sido, gu_hint)
            except Exception as exc:
                print(f"  [error] 4th {etype} {fname}: {exc}")
    print(f"  After 4th: {len(ALL_ROWS)} rows")


# ─────────────────────────────────────────────────────────────────────────────
# 5회 processing
# ─────────────────────────────────────────────────────────────────────────────
def process_5th():
    base = os.path.join(BASE, "전국동시지방선거 개표결과(제3회~제6회)", "제5회 전국동시지방선거 개표자료")
    type_dir_map = {
        "01_시도지사": "시도지사", "02_구시군장": "구시군장",
        "03_시도의원": "시도의원", "04_광역의원비례대표": "광역비례",
        "05_구시군의원": "구시군의원", "06_기초의원비례대표": "기초비례",
        "07_교육감": "교육감", "08_교육의원": "교육의원",
    }
    for dirname, etype in type_dir_map.items():
        dirpath = os.path.join(base, dirname)
        if not os.path.isdir(dirpath):
            continue
        for fpath in sorted(find_xls_files(dirpath)):
            fname = os.path.basename(fpath)
            # sido hint: try from the file path components
            path_parts = fpath.replace(dirpath, "").strip(os.sep).split(os.sep)
            sido_hint = ""
            for part in path_parts:
                candidate = re.sub(r"^\d+_", "", os.path.splitext(part)[0])
                # Check if it looks like a sido name
                if re.search(r"(광역시|특별시|도|특별자치)", candidate):
                    sido_hint = candidate
                    break
            try:
                parse_5th_6th_file(fpath, 5, etype, sido_hint)
            except Exception as exc:
                print(f"  [error] 5th {etype} {fname}: {exc}")
    print(f"  After 5th: {len(ALL_ROWS)} rows")


# ─────────────────────────────────────────────────────────────────────────────
# 6회 processing
# ─────────────────────────────────────────────────────────────────────────────
def process_6th():
    base = os.path.join(BASE, "전국동시지방선거 개표결과(제3회~제6회)", "제6회 전국동시지방선거 개표결과")

    # Sejon special (multi-sheet xlsx)
    sejon_path = os.path.join(base, "제6회 전국동시지방선거 읍면동별 개표자료(세종).xlsx")
    if os.path.exists(sejon_path):
        try:
            wb = openpyxl.load_workbook(sejon_path, data_only=True)
            sheet_etype_map = {
                "세종시장선거": "구시군장", "세종시교육감선거": "교육감",
                "세종시의원선거": "시도의원", "비례대표세종시의원선거": "광역비례",
            }
            for sheet_name, etype in sheet_etype_map.items():
                if sheet_name in wb.sheetnames:
                    # Save as temp and parse
                    ws = wb[sheet_name]
                    # Process inline
                    all_rows = list(ws.iter_rows(values_only=True))
                    _parse_6th_세종_sheet(all_rows, 6, etype, "세종특별자치시")
        except Exception as exc:
            print(f"  [error] 6th 세종: {exc}")

    # All other regions
    for region_dirname in sorted(os.listdir(base)):
        region_path = os.path.join(base, region_dirname)
        if not os.path.isdir(region_path):
            continue

        # Extract sido from dirname
        match = re.search(r"\((.+)\)$", region_dirname)
        sido_hint = match.group(1) if match else ""

        # Map short names to full names
        sido_full_map = {
            "강원": "강원도", "경기": "경기도", "경남": "경상남도", "경북": "경상북도",
            "광주": "광주광역시", "대구": "대구광역시", "대전": "대전광역시",
            "부산": "부산광역시", "서울": "서울특별시", "세종": "세종특별자치시",
            "울산": "울산광역시", "인천": "인천광역시", "전남": "전라남도",
            "전북": "전라북도", "제주": "제주특별자치도", "충남": "충청남도",
            "충북": "충청북도",
        }
        sido_full = sido_full_map.get(sido_hint, sido_hint)

        # Find all xls/xlsx under this region
        all_files = find_xls_files(region_path)
        for fpath in sorted(all_files):
            etype = infer_election_type_from_path(fpath)
            try:
                parse_5th_6th_file(fpath, 6, etype, sido_full)
            except Exception as exc:
                print(f"  [error] 6th {region_dirname} {os.path.basename(fpath)}: {exc}")

    print(f"  After 6th: {len(ALL_ROWS)} rows")


def _parse_6th_세종_sheet(all_rows, round_num, election_type, sido):
    """Parse a single sheet from the 세종 multi-sheet xlsx."""
    nrows = len(all_rows)
    ncols = max(len(row) for row in all_rows) if all_rows else 0

    # Find header row
    header_row_idx = None
    for idx, row in enumerate(all_rows[:6]):
        row_strs = [str(v or "").strip() for v in row]
        if any(v in ("읍면동명", "선거인수", "투표수") for v in row_strs):
            header_row_idx = idx
            break

    if header_row_idx is None:
        return

    header = [str(all_rows[header_row_idx][col] or "").strip()
              if col < len(all_rows[header_row_idx]) else ""
              for col in range(ncols)]
    cand_row = [str(all_rows[header_row_idx + 1][col] or "").strip()
                if col < len(all_rows[header_row_idx + 1]) else ""
                for col in range(ncols)] if header_row_idx + 1 < nrows else []

    emd_col = voters_col = votes_col = invalid_col = abstain_col = total_col = None
    for col_idx, hdr in enumerate(header):
        if hdr in ("읍면동명",):
            emd_col = col_idx
        elif "선거인수" in hdr:
            voters_col = col_idx
        elif hdr == "투표수":
            votes_col = col_idx
        elif "무효" in hdr:
            invalid_col = col_idx
        elif "기권" in hdr:
            abstain_col = col_idx
        elif hdr == "계":
            total_col = col_idx

    if emd_col is None:
        emd_col = 0
    if voters_col is None:
        voters_col = 2
    if votes_col is None:
        votes_col = 3

    start_col = votes_col + 1
    end_col = total_col or invalid_col or ncols
    cand_cols = [(col_idx, cand_row[col_idx])
                 for col_idx in range(start_col, end_col)
                 if col_idx < len(cand_row) and cand_row[col_idx] and cand_row[col_idx] != "계"]

    data_start = header_row_idx + 2
    while data_start < nrows:
        row = all_rows[data_start]
        if any(v is not None and str(v).strip() not in ("", "0") for v in row[:5]):
            break
        data_start += 1

    for row in all_rows[data_start:]:
        vals = list(row)
        while len(vals) < ncols:
            vals.append(None)

        emd = str(vals[emd_col] or "").strip()
        if not emd or is_summary_row(emd):
            continue
        # 구분 컬럼으로 level 결정
        row_level_sejong = "당일투표"
        if emd_col + 1 < len(vals):
            구분 = str(vals[emd_col + 1] or "").strip()
            if 구분 in ("관내", "관외"):
                continue
            row_level_sejong = get_level(구분) if 구분 else get_level(emd)

        voters = vals[voters_col] if voters_col < len(vals) else None
        votes = vals[votes_col] if votes_col < len(vals) else None
        invalid = vals[invalid_col] if invalid_col is not None and invalid_col < len(vals) else None
        abstain = vals[abstain_col] if abstain_col is not None and abstain_col < len(vals) else None

        for col_idx, cand_header in cand_cols:
            if col_idx >= len(vals):
                continue
            score = vals[col_idx]
            party, candidate = split_party_name(cand_header)
            if clean_num(score) is None:
                continue
            emit_row(round_num, election_type, sido, "세종특별자치시", emd,
                     "세종특별자치시", voters, votes, candidate, party, score, invalid, abstain,
                     level=row_level_sejong)


# ─────────────────────────────────────────────────────────────────────────────
# 7회 processing
# ─────────────────────────────────────────────────────────────────────────────
def process_7th():
    base = os.path.join(BASE, "전국동시지방선거 개표결과(제7회)")
    file_etype_map = {
        "01-(시도지사)": "시도지사", "02-(구시군의장)": "구시군장",
        "03-(시도의회의원)": "시도의원", "04-(구시군의회의원)": "구시군의원",
        "05-(광역비례)": "광역비례", "06-(기초비례)": "기초비례",
        "07-(교육감)": "교육감", "08-(교육의원)": "교육의원",
        "09-(국회의원재보궐)": "국회의원재보궐",
    }
    for fname in sorted(os.listdir(base)):
        if not fname.endswith(".xlsx"):
            continue
        etype = "기타"
        fname_nfc = unicodedata.normalize("NFC", fname)
        for key, val in file_etype_map.items():
            if key in fname_nfc:
                etype = val
                break
        fpath = os.path.join(base, fname)
        try:
            parse_7th_xlsx(fpath, etype)
        except Exception as exc:
            print(f"  [error] 7th {fname}: {exc}")
    print(f"  After 7th: {len(ALL_ROWS)} rows")


# ─────────────────────────────────────────────────────────────────────────────
# 8회 processing
# ─────────────────────────────────────────────────────────────────────────────
def process_8th():
    base = os.path.join(BASE, "제8회_전국동시지방선거_읍면동별_개표결과-게시판게시")
    file_etype_map = [
        ("시도지사", "시도지사"), ("구시군장", "구시군장"),
        ("시도의원", "시도의원"), ("구시군의원", "구시군의원"),
        ("광역비례", "광역비례"), ("기초비례", "기초비례"),
        ("교육감", "교육감"), ("교육의원", "교육의원"),
        ("국회의원", "국회의원재보궐"),
    ]
    for fname in sorted(os.listdir(base)):
        if not fname.endswith(".xlsx"):
            continue
        etype = "기타"
        # Normalize to NFC to handle macOS NFD filenames
        base_name = unicodedata.normalize("NFC", os.path.splitext(fname)[0])
        for key, val in file_etype_map:
            if key in base_name:
                etype = val
                break
        fpath = os.path.join(base, fname)
        try:
            parse_8th_xlsx(fpath, etype)
        except Exception as exc:
            print(f"  [error] 8th {fname}: {exc}")
    print(f"  After 8th: {len(ALL_ROWS)} rows")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Processing 3rd election...")
    process_3rd()

    print("Processing 4th election...")
    process_4th()

    print("Processing 5th election...")
    process_5th()

    print("Processing 6th election...")
    process_6th()

    print("Processing 7th election...")
    process_7th()

    print("Processing 8th election...")
    process_8th()

    print(f"\nTotal rows before dedup: {len(ALL_ROWS)}")

    df = pd.DataFrame(ALL_ROWS, columns=FINAL_COLS)

    # Remove duplicate rows
    df = df.drop_duplicates()

    # Remove summary/artifact rows
    df = df[~df["읍면동"].str.contains("합  계|잘못", na=False)]
    df = df[~df["구시군"].str.contains("잘못|위원회명", na=False)]
    df = df[df["읍면동"].str.strip() != "위원회명"]
    # Remove rows with 합 계 (single space) summary
    df = df[df["읍면동"].str.strip() != "합 계"]
    # Remove rows where 후보자 looks like a header artifact
    df = df[~df["후보자"].str.contains("득 표|후보자별|별 득 표", na=False)]

    # Type conversion
    for col in ["선거인수", "투표수", "득표수", "무효투표수", "기권수"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    out_path = os.path.join(OUT_DIR, "지방선거.csv")
    df.to_csv(out_path, index=False, encoding="utf-8-sig")

    print(f"\n=== 완료 ===")
    print(f"저장 위치: {out_path}")
    print(f"전체 행 수: {len(df):,}")
    print(f"컬럼: {list(df.columns)}")
    print(f"\n--- 회차별 행 수 ---")
    print(df.groupby(["선거_회차", "선거종류"]).size().to_string())
    print(f"\n--- 샘플 5행 ---")
    print(df.head(5).to_string())
